import asyncio
from datetime import datetime, timedelta
from aiogram import Router, F
from aiogram.types import CallbackQuery, Message, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from handlers import is_event_hidden
from keyboards import get_to_main_keyboard, get_event_notification_keyboard

import config
from database.db import async_session
from database.models import User, Event, Raffle, Registration, Admin, FeedbackMessage, PartnerEvent, EventSeries, SeriesQuestion, SeriesEvent, SeriesApplication
from admin_keyboards import (
    get_admin_main_keyboard,
    get_skip_keyboard,
    get_cancel_keyboard,
    get_address_keyboard,
    get_admin_events_keyboard,
    get_event_select_edit_keyboard,
    get_event_sections_keyboard,
    get_after_edit_event_keyboard,
    get_event_delete_keyboard,
    get_confirm_delete_event_keyboard,
    get_after_deleted_event_keyboard,
    get_admin_tags_keyboard,
    get_after_tag_created_keyboard,
    get_tag_delete_keyboard,
    get_after_tag_deleted_keyboard,
    get_admin_raffles_keyboard,
    get_raffle_select_edit_keyboard,
    get_raffle_sections_keyboard,
    get_after_edit_raffle_keyboard,
    get_raffle_delete_keyboard,
    get_confirm_delete_raffle_keyboard,
    get_after_deleted_raffle_keyboard,
    get_raffle_link_choice_keyboard,
    get_raffle_link_event_keyboard,
    get_after_raffle_save_keyboard,
    get_admin_rights_keyboard,
    get_admin_tags_selection_keyboard,
    get_image_add_keyboard,
    get_after_event_created_keyboard,
    get_finish_event_keyboard,
    get_admin_post_mats_keyboard,
    get_admin_post_mats_event_select_keyboard,
    get_after_post_mats_saved_keyboard,
    get_confirm_del_post_mats_keyboard,
    get_post_mats_tags_keyboard,
    get_admin_delete_select_keyboard,
    get_raffle_edit_link_event_keyboard,
    get_admin_archive_tags_keyboard,
    get_admin_archive_events_keyboard,
    get_admin_export_type_keyboard,
    get_admin_export_events_keyboard,
    get_admin_export_archive_tags_keyboard,
    get_admin_export_archive_events_keyboard,
    get_admin_export_back_keyboard,
    get_feedback_navigation_keyboard,
    get_admin_partner_events_keyboard,
    get_admin_series_list_keyboard,
    get_admin_series_actions_keyboard,
    get_questionnaire_loop_keyboard,
    get_series_events_select_keyboard,
    get_winners_checkbox_keyboard,
    get_winners_confirm_keyboard,
    get_winners_dispatch_keyboard
)

router = Router()

# ==========================================
# FSM ОПРЕДЕЛЕНИЯ СОСТОЯНИЙ
# ==========================================

class EventForm(StatesGroup):
    title = State()
    title_url = State()
    date = State()
    time = State()
    address = State()
    description = State()
    reg_url = State()
    stream_url = State()
    tags = State()
    image = State()
    hide_date = State()
    hide_time = State()


class EditEventForm(StatesGroup):
    event_id = State()
    section = State() # Название поля, которое меняем
    value = State()


class TagForm(StatesGroup):
    name = State()


class RaffleForm(StatesGroup):
    title = State()
    description = State()
    url = State()
    hide_date = State()
    hide_time = State()
    event_id = State()


class PartnerEventForm(StatesGroup):
    title = State()
    date = State()
    description = State()
    link = State()


class EditRaffleForm(StatesGroup):
    raffle_id = State()
    section = State()
    value = State()


class AdminRightsForm(StatesGroup):
    username = State()


class PostMatsForm(StatesGroup):
    event_id = State()
    photos_url = State()
    photographer_name = State()
    photographer_url = State()
    stream_record_url = State()
    article_url = State()
    presentations_url = State()
    other_materials_url = State()


class CreateSeriesForm(StatesGroup):
    title = State()
    description = State()
    image = State()


class AddSeriesEventForm(StatesGroup):
    series_id = State()
    topic = State()
    date = State()
    time = State()
    extra_text = State()


class SeriesQuestionnaireForm(StatesGroup):
    series_id = State()
    asking_question = State()


class SelectWinnersForm(StatesGroup):
    series_id = State()
    series_event_id = State()
    selected_app_ids = State() # list/set of app_ids
    confirming = State()
    entering_extra_text = State()



# ==========================================
# ПРОВЕРКА ДОСТУПА
# ==========================================

async def is_user_admin(username: str | None, session: AsyncSession) -> bool:
    if not username:
        return False
    if config.is_super_admin(username):
        return True
    clean_username = username.lstrip('@').lower()
    query = select(Admin).where(func.lower(Admin.username) == clean_username)
    result = await session.execute(query)
    return result.scalar_one_or_none() is not None


async def get_admin_welcome_text(session: AsyncSession) -> str:
    total_query = select(func.count(User.telegram_id))
    total_res = await session.execute(total_query)
    total_users = total_res.scalar() or 0

    now = datetime.now()
    last_24h = (now - timedelta(hours=24)).strftime("%Y-%m-%d %H:%M:%S")
    last_7d = (now - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")

    day_query = select(func.count(User.telegram_id)).where(User.created_at >= last_24h)
    day_res = await session.execute(day_query)
    new_users_day = day_res.scalar() or 0

    week_query = select(func.count(User.telegram_id)).where(User.created_at >= last_7d)
    week_res = await session.execute(week_query)
    new_users_week = week_res.scalar() or 0

    return (
        "Привет, Админ!\n\n"
        f"Уникальных пользователей бота: {total_users}\n"
        f"Новых пользователей за сутки: {new_users_day}\n"
        f"Новых пользователей за неделю: {new_users_week}\n\n"
        "Что вы хотите сделать?"
    )


@router.callback_query(F.data == "btn_admin")
async def process_admin_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        welcome_text = await get_admin_welcome_text(session)
        await callback.message.answer(
            welcome_text,
            reply_markup=get_admin_main_keyboard()
        )
    await callback.answer()


@router.callback_query(F.data == "admin_cancel")
async def process_admin_cancel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    async with async_session() as session:
        welcome_text = await get_admin_welcome_text(session)
        await callback.message.answer(
            welcome_text,
            reply_markup=get_admin_main_keyboard()
        )
    await callback.answer()


# ==========================================
# 2. CRUD МЕРОПРИЯТИЙ
# ==========================================

@router.callback_query(F.data == "admin_edit_events")
async def process_admin_edit_events(callback: CallbackQuery):
    await callback.message.answer(
        "Редактируем мероприятие. Выберите действие.",
        reply_markup=get_admin_events_keyboard()
    )
    await callback.answer()


# ----- Добавление мероприятия -----

@router.callback_query(F.data == "admin_add_event")
async def process_admin_add_event(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.title)
    await callback.message.answer(
        "Вопрос 1: Название",
        reply_markup=get_skip_keyboard("skip_title")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_title", EventForm.title)
async def process_skip_title(callback: CallbackQuery, state: FSMContext):
    await state.update_data(title="")
    await state.set_state(EventForm.title_url)
    await callback.message.answer(
        "Вопрос 2: Ссылка, вшиваемая в название",
        reply_markup=get_skip_keyboard("skip_title_url", show_back=True, back_callback="back_title_url")
    )
    await callback.answer()


@router.message(EventForm.title)
async def process_add_event_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await state.set_state(EventForm.title_url)
    await message.answer(
        "Вопрос 2: Ссылка, вшиваемая в название",
        reply_markup=get_skip_keyboard("skip_title_url", show_back=True, back_callback="back_title_url")
    )


@router.callback_query(F.data == "back_title_url", EventForm.title_url)
async def process_back_title_url(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.title)
    await callback.message.answer(
        "Вопрос 1: Название",
        reply_markup=get_skip_keyboard("skip_title")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_title_url", EventForm.title_url)
async def process_skip_title_url(callback: CallbackQuery, state: FSMContext):
    await state.update_data(title_url=None)
    await state.set_state(EventForm.date)
    await callback.message.answer(
        "Вопрос 3: Отображаемая дата мероприятия (в формате ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_date")
    )
    await callback.answer()


@router.message(EventForm.title_url)
async def process_add_event_title_url(message: Message, state: FSMContext):
    await state.update_data(title_url=message.text)
    await state.set_state(EventForm.date)
    await message.answer(
        "Вопрос 3: Отображаемая дата мероприятия (в формате ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_date")
    )


@router.callback_query(F.data == "back_date", EventForm.date)
async def process_back_date(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.title_url)
    await callback.message.answer(
        "Вопрос 2: Ссылка, вшиваемая в название",
        reply_markup=get_skip_keyboard("skip_title_url", show_back=True, back_callback="back_title_url")
    )
    await callback.answer()


def validate_date_format(date_str: str) -> bool:
    date_str = date_str.strip()
    if "-" in date_str:
        parts = date_str.split("-")
        if len(parts) == 2:
            try:
                datetime.strptime(parts[0].strip(), "%d.%m.%Y")
                datetime.strptime(parts[1].strip(), "%d.%m.%Y")
                return True
            except ValueError:
                return False
        return False
    else:
        try:
            datetime.strptime(date_str, "%d.%m.%Y")
            return True
        except ValueError:
            return False


@router.message(EventForm.date)
async def process_add_event_date(message: Message, state: FSMContext):
    if not validate_date_format(message.text):
        await message.answer(
            "Неверный формат даты, попробуйте еще раз. Ожидается формат ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_date")
        )
        return
        
    await state.update_data(date=message.text.strip())
    await state.set_state(EventForm.time)
    await message.answer(
        "Вопрос 4: Отображаемое время мероприятия (строго в формате ЧЧ:ММ-ЧЧ:ММ, например 19:00-21:00)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_time")
    )


@router.callback_query(F.data == "back_time", EventForm.time)
async def process_back_time(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.date)
    await callback.message.answer(
        "Вопрос 3: Отображаемая дата мероприятия (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_date")
    )
    await callback.answer()


@router.message(EventForm.time)
async def process_add_event_time(message: Message, state: FSMContext):
    t_str = message.text.strip()
    normalized_t = t_str.replace("—", "-").replace(" - ", "-").replace(" -", "-").replace("- ", "-")
    parts = normalized_t.split("-")
    
    valid = False
    if len(parts) == 2:
        try:
            datetime.strptime(parts[0].strip(), "%H:%M")
            datetime.strptime(parts[1].strip(), "%H:%M")
            valid = True
        except ValueError:
            pass
            
    if not valid:
        await message.answer(
            "Неверный формат времени, попробуйте еще раз. Ожидается формат ЧЧ:ММ-ЧЧ:ММ (например, 19:00-21:00)",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_time")
        )
        return

    await state.update_data(time=t_str)
    await state.set_state(EventForm.address)
    await message.answer(
        "Вопрос 5: Отображаемый адрес мероприятия",
        reply_markup=get_address_keyboard(show_back=True, back_callback="back_address")
    )


@router.callback_query(F.data == "back_address", EventForm.address)
async def process_back_address(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.time)
    await callback.message.answer(
        "Вопрос 4: Отображаемое время мероприятия (строго в формате ЧЧ:ММ-ЧЧ:ММ, например 19:00-21:00)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_time")
    )
    await callback.answer()


@router.callback_query(F.data == "admin_event_default_address", EventForm.address)
async def process_add_event_default_address(callback: CallbackQuery, state: FSMContext):
    await state.update_data(address="Москва, Пантелеевская 53")
    await state.set_state(EventForm.description)
    await callback.message.answer(
        "Вопрос 6: Основной текст",
        reply_markup=get_skip_keyboard("skip_desc", show_back=True, back_callback="back_desc")
    )
    await callback.answer()


@router.message(EventForm.address)
async def process_add_event_address(message: Message, state: FSMContext):
    await state.update_data(address=message.text)
    await state.set_state(EventForm.description)
    await message.answer(
        "Вопрос 6: Основной текст",
        reply_markup=get_skip_keyboard("skip_desc", show_back=True, back_callback="back_desc")
    )


@router.callback_query(F.data == "back_desc", EventForm.description)
async def process_back_desc(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.address)
    await callback.message.answer(
        "Вопрос 5: Отображаемый адрес мероприятия",
        reply_markup=get_address_keyboard(show_back=True, back_callback="back_address")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_desc", EventForm.description)
async def process_skip_desc(callback: CallbackQuery, state: FSMContext):
    await state.update_data(description=None)
    await state.set_state(EventForm.reg_url)
    await callback.message.answer(
        "Вопрос 7: Ссылка на регистрацию",
        reply_markup=get_skip_keyboard("skip_reg_url", show_back=True, back_callback="back_reg_url")
    )
    await callback.answer()


@router.message(EventForm.description)
async def process_add_event_desc(message: Message, state: FSMContext):
    await state.update_data(description=message.html_text)
    await state.set_state(EventForm.reg_url)
    await message.answer(
        "Вопрос 7: Ссылка на регистрацию",
        reply_markup=get_skip_keyboard("skip_reg_url", show_back=True, back_callback="back_reg_url")
    )


@router.callback_query(F.data == "back_reg_url", EventForm.reg_url)
async def process_back_reg_url(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.description)
    await callback.message.answer(
        "Вопрос 6: Основной текст",
        reply_markup=get_skip_keyboard("skip_desc", show_back=True, back_callback="back_desc")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_reg_url", EventForm.reg_url)
async def process_skip_reg_url(callback: CallbackQuery, state: FSMContext):
    await state.update_data(reg_url=None)
    await state.set_state(EventForm.stream_url)
    await callback.message.answer(
        "Вопрос 8: Ссылка на трансляцию",
        reply_markup=get_skip_keyboard("skip_stream_url", show_back=True, back_callback="back_stream_url")
    )
    await callback.answer()


@router.message(EventForm.reg_url)
async def process_add_event_reg_url(message: Message, state: FSMContext):
    await state.update_data(reg_url=message.text)
    await state.set_state(EventForm.stream_url)
    await message.answer(
        "Вопрос 8: Ссылка на трансляцию",
        reply_markup=get_skip_keyboard("skip_stream_url", show_back=True, back_callback="back_stream_url")
    )


@router.callback_query(F.data == "back_stream_url", EventForm.stream_url)
async def process_back_stream_url(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.reg_url)
    await callback.message.answer(
        "Вопрос 7: Ссылка на регистрацию",
        reply_markup=get_skip_keyboard("skip_reg_url", show_back=True, back_callback="back_reg_url")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_stream_url", EventForm.stream_url)
async def process_skip_stream_url(callback: CallbackQuery, state: FSMContext):
    await state.update_data(stream_url=None)
    await state.set_state(EventForm.tags)
    await state.update_data(selected_tags=[])
    await callback.message.answer(
        "Вопрос 9: Выбор тегов мероприятия",
        reply_markup=get_admin_tags_selection_keyboard(config.DEFAULT_TAGS, [], step_mode=True)
    )
    await callback.answer()


@router.message(EventForm.stream_url)
async def process_add_event_stream_url(message: Message, state: FSMContext):
    await state.update_data(stream_url=message.text)
    await state.set_state(EventForm.tags)
    await state.update_data(selected_tags=[])
    await message.answer(
        "Вопрос 9: Выбор тегов мероприятия",
        reply_markup=get_admin_tags_selection_keyboard(config.DEFAULT_TAGS, [], step_mode=True)
    )


@router.callback_query(F.data == "back_tags", EventForm.tags)
async def process_back_tags(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.stream_url)
    await callback.message.answer(
        "Вопрос 8: Ссылка на трансляцию",
        reply_markup=get_skip_keyboard("skip_stream_url", show_back=True, back_callback="back_stream_url")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_tags", EventForm.tags)
async def process_skip_tags(callback: CallbackQuery, state: FSMContext):
    await state.update_data(selected_tags=[])
    await state.set_state(EventForm.image)
    await state.update_data(temp_images=[])
    await callback.message.answer(
        "Вопрос 10: Добавление изображения без сжатия (файлом)",
        reply_markup=get_image_add_keyboard(has_images=False)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_evtag_"), EventForm.tags)
async def process_add_event_tags_toggle(callback: CallbackQuery, state: FSMContext):
    action = callback.data.split("_")[2]
    
    if action == "confirm":
        state_data = await state.get_data()
        selected = state_data.get("selected_tags", [])
        await state.set_state(EventForm.image)
        await state.update_data(temp_images=[])
        await callback.message.answer(
            "Вопрос 10: Добавление изображения без сжатия (файлом)",
            reply_markup=get_image_add_keyboard(has_images=False)
        )
        await callback.answer()
        return
        
    try:
        idx = int(action)
        tag_name = config.DEFAULT_TAGS[idx]
    except (ValueError, IndexError):
        tag_name = action

    state_data = await state.get_data()
    selected = list(state_data.get("selected_tags", []))
    if tag_name in selected:
        selected.remove(tag_name)
    else:
        selected.append(tag_name)
        
    await state.update_data(selected_tags=selected)
    await callback.message.edit_reply_markup(
        reply_markup=get_admin_tags_selection_keyboard(config.DEFAULT_TAGS, selected, step_mode=True)
    )
    await callback.answer()


@router.callback_query(F.data == "back_image", EventForm.image)
async def process_back_image(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.tags)
    state_data = await state.get_data()
    selected = state_data.get("selected_tags", [])
    await callback.message.answer(
        "Вопрос 9: Выбор тегов мероприятия",
        reply_markup=get_admin_tags_selection_keyboard(config.DEFAULT_TAGS, selected, step_mode=True)
    )
    await callback.answer()


@router.message(EventForm.image)
async def process_add_event_image(message: Message, state: FSMContext):
    import os
    from aiogram.types import FSInputFile

    # ПРОВЕРКА: Картинка должна быть отправлена ИСКЛЮЧИТЕЛЬНО как файл (документ)
    if not message.document:
        await message.answer(
            "Ошибка: изображение должно быть отправлено как файл (без сжатия). Пожалуйста, попробуйте еще раз:",
            reply_markup=get_image_add_keyboard(has_images=False)
        )
        return
        
    # Проверяем mimetype (это должен быть image)
    mime = message.document.mime_type or ""
    if not mime.startswith("image/"):
        await message.answer(
            "Ошибка: загруженный файл не является картинкой. Пожалуйста, попробуйте еще раз:",
            reply_markup=get_image_add_keyboard(has_images=False)
        )
        return

    doc_file_id = message.document.file_id
    
    # Конвертируем документ в фото на стороне ТГ, отправив его как фото
    file_info = await message.bot.get_file(doc_file_id)
    temp_path = f"temp_{doc_file_id}.jpg"
    await message.bot.download_file(file_info.file_path, temp_path)
    
    # Отправляем фото и забираем file_id фото
    sent_msg = await message.answer_photo(photo=FSInputFile(temp_path), caption="Загрузка изображения...")
    photo_file_id = sent_msg.photo[-1].file_id
    await sent_msg.delete()
    
    if os.path.exists(temp_path):
        os.remove(temp_path)

    state_data = await state.get_data()
    images = list(state_data.get("temp_images", []))
    images.append(photo_file_id)
    await state.update_data(temp_images=images)
    
    await message.answer(
        f"Изображение успешно получено! Всего загружено картинок: {len(images)}",
        reply_markup=get_image_add_keyboard(has_images=True)
    )


@router.callback_query(F.data == "admin_image_add_more", EventForm.image)
async def process_image_add_more(callback: CallbackQuery):
    await callback.message.answer(
        "Вопрос 10: Добавление изображения без сжатия (файлом)",
        reply_markup=get_image_add_keyboard(has_images=True)
    )
    await callback.answer()


@router.callback_query(F.data == "skip_image", EventForm.image)
async def process_skip_image_btn(callback: CallbackQuery, state: FSMContext):
    await state.update_data(images=[])
    await state.set_state(EventForm.hide_date)
    await callback.message.answer(
        "Вопрос 11: Дата, когда мероприятие должно исчезнуть из видимости (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_skip_keyboard("skip_hide_date", show_back=True, back_callback="back_hide_date")
    )
    await callback.answer()


@router.callback_query(F.data == "admin_image_save", EventForm.image)
async def process_image_save(callback: CallbackQuery, state: FSMContext):
    state_data = await state.get_data()
    images = state_data.get("temp_images", [])
    await state.update_data(images=images)
    await state.set_state(EventForm.hide_date)
    await callback.message.answer(
        "Вопрос 11: Дата, когда мероприятие должно исчезнуть из видимости (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_skip_keyboard("skip_hide_date", show_back=True, back_callback="back_hide_date")
    )
    await callback.answer()


@router.callback_query(F.data == "back_hide_date", EventForm.hide_date)
async def process_back_hide_date(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.image)
    state_data = await state.get_data()
    images = state_data.get("temp_images", [])
    await callback.message.answer(
        "Вопрос 10: Добавление изображения без сжатия (файлом)",
        reply_markup=get_image_add_keyboard(has_images=len(images) > 0)
    )
    await callback.answer()


@router.callback_query(F.data == "skip_hide_date", EventForm.hide_date)
async def process_skip_hide_date(callback: CallbackQuery, state: FSMContext):
    await state.update_data(hide_date=None)
    await state.update_data(hide_time=None)
    await state.set_state(EventForm.hide_time) # перейдем в стейт тайм, чтобы кнопка финиша работала
    await callback.message.answer(
        "Вопрос 12: Время, когда мероприятие должно исчезнуть из видимости (строго в формате ЧЧ:ММ, формат 24 часа)",
        reply_markup=get_finish_event_keyboard()
    )
    await callback.answer()


@router.message(EventForm.hide_date)
async def process_add_event_hide_date(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text.strip(), "%d.%m.%Y")
    except ValueError:
        await message.answer(
            "Неверный формат даты, попробуйте еще раз",
            reply_markup=get_skip_keyboard("skip_hide_date", show_back=True, back_callback="back_hide_date")
        )
        return

    await state.update_data(hide_date=message.text.strip())
    await state.set_state(EventForm.hide_time)
    await message.answer(
        "Вопрос 12: Время, когда мероприятие должно исчезнуть из видимости (строго в формате ЧЧ:ММ, формат 24 часа)",
        reply_markup=get_finish_event_keyboard()
    )


@router.callback_query(F.data == "back_hide_time", EventForm.hide_time)
async def process_back_hide_time(callback: CallbackQuery, state: FSMContext):
    await state.set_state(EventForm.hide_date)
    await callback.message.answer(
        "Вопрос 11: Дата, когда мероприятие должно исчезнуть из видимости (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_skip_keyboard("skip_hide_date", show_back=True, back_callback="back_hide_date")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_hide_time", EventForm.hide_time)
async def process_skip_hide_time_btn(callback: CallbackQuery, state: FSMContext):
    await state.update_data(hide_time=None)
    await save_event_to_db(callback.message, state)
    await callback.answer()


@router.message(EventForm.hide_time)
async def process_add_event_hide_time(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text.strip(), "%H:%M")
    except ValueError:
        await message.answer(
            "Неверный формат времени, попробуйте еще раз",
            reply_markup=get_finish_event_keyboard()
        )
        return

    await state.update_data(hide_time=message.text.strip())
    await save_event_to_db(message, state)


@router.callback_query(F.data == "admin_finish_event", EventForm.hide_time)
async def process_finish_event_btn(callback: CallbackQuery, state: FSMContext):
    await save_event_to_db(callback.message, state)
    await callback.answer()


async def save_event_to_db(message: Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()

    async with async_session() as session:
        new_event = Event(
            title=data.get("title", ""),
            title_url=data.get("title_url"),
            date=data.get("date", ""),
            time=data.get("time", ""),
            address=data.get("address", ""),
            description=data.get("description"),
            reg_url=data.get("reg_url"),
            stream_url=data.get("stream_url"),
            tags=data.get("selected_tags", []),
            images=data.get("images", []),
            hide_date=data.get("hide_date"),
            hide_time=data.get("hide_time")
        )
        session.add(new_event)
        await session.commit()
        
        if message.bot:
            asyncio.create_task(send_event_creation_notifications(message.bot, new_event))
        
        # Получаем данные заново для отображения карточки
        event_id = new_event.id

    async with async_session() as session:
        event = await session.get(Event, event_id)
        title_html = f"<b>{event.title}</b>"
        if event.title_url:
            title_html = f'<b><a href="{event.title_url}">{event.title}</a></b>'
            
        tags_str = " ".join([f"#{tag}" for tag in event.tags])
        desc = event.description or ""
        
        links_list = []
        if event.reg_url:
            links_list.append(f'→ <a href="{event.reg_url}">Регистрация на сайте</a>\nИли в телеграм-боте 👇')
        if event.stream_url:
            links_list.append(f'→ <a href="{event.stream_url}">Трансляция</a>')
        links_str = "\n\n" + "\n".join(links_list) if links_list else ""
        
        text = (
            f"Мероприятие добавлено!\n\n"
            f"{title_html}\n\n"
            f"{desc}\n\n"
            f"📆 <b>Дата:</b> {config.format_display_date(event.date)}\n"
            f"⏳ <b>Время:</b> {event.time}\n"
            f"📍 <b>Место:</b> {event.address}"
            f"{links_str}\n\n"
            f"{tags_str}"
        )
        
        kb = get_after_event_created_keyboard()
        
        if event.images and len(event.images) > 0:
            await message.answer_photo(
                photo=event.images[0],
                caption=text,
                reply_markup=kb,
                parse_mode="HTML"
            )
        else:
            await message.answer(
                text,
                reply_markup=kb,
                parse_mode="HTML",
                disable_web_page_preview=True
            )


# ----- Редактирование существующего мероприятия -----

@router.callback_query(F.data == "admin_edit_event_list")
async def process_admin_edit_event_list(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        
        active_events = [e for e in events if not is_event_hidden(e)]
        
        await callback.message.answer(
            "Какое мероприятие редактировать?",
            reply_markup=get_event_select_edit_keyboard(active_events)
        )
    await callback.answer()


@router.callback_query(F.data == "admin_edit_archive_tags")
async def process_admin_edit_archive_tags(callback: CallbackQuery):
    await callback.message.answer(
        "Выберите тему для поиска архивных мероприятий:",
        reply_markup=get_admin_archive_tags_keyboard(config.DEFAULT_TAGS)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_archtag_"))
async def process_admin_archtag_select(callback: CallbackQuery):
    tag_idx = int(callback.data.split("_")[2])
    try:
        tag = config.DEFAULT_TAGS[tag_idx]
    except IndexError:
        await callback.answer("Тема не найдена.")
        return

    async with async_session() as session:
        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        
        archived_events = [e for e in events if is_event_hidden(e) and tag in (e.tags or [])]
        
        if not archived_events:
            await callback.message.answer(
                f"Нет архивных мероприятий по теме «{tag}».",
                reply_markup=get_admin_archive_tags_keyboard(config.DEFAULT_TAGS)
            )
        else:
            await callback.message.answer(
                f"Архивные мероприятия по теме «{tag}»:\nВыберите мероприятие для редактирования.",
                reply_markup=get_admin_archive_events_keyboard(archived_events, tag_idx)
            )
    await callback.answer()



@router.callback_query(F.data.startswith("admin_select_edit_event_"))
async def process_admin_select_edit_event(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[4])
    await callback.message.answer(
        "Какой раздел редактировать?",
        reply_markup=get_event_sections_keyboard(event_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("edit_evsec_"))
async def process_edit_evsec_start(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    event_id = int(parts[2])
    section = parts[3]
    
    await state.set_state(EditEventForm.value)
    await state.update_data(event_id=event_id, section=section)
    
    section_titles = {
        "title": "Название",
        "titleurl": "Ссылка, вшиваемая в название",
        "date": "Отображаемая дата мероприятия (ДД.ММ.ГГГГ)",
        "time": "Отображаемое время мероприятия (строго в формате ЧЧ:ММ-ЧЧ:ММ, например 19:00-21:00)",
        "address": "Отображаемый адрес мероприятия",
        "desc": "Основной текст",
        "regurl": "Ссылка на регистрацию",
        "streamurl": "Ссылка на трансляцию",
        "tags": "Выбор тегов",
        "image": "Добавление изображения (файлом без сжатия)",
        "hidedate": "Дата скрытия (ДД.ММ.ГГГГ)",
        "hidetime": "Время скрытия (ЧЧ:ММ)"
    }
    
    title = section_titles.get(section, "значение")
    
    if section == "tags":
        async with async_session() as session:
            event = await session.get(Event, event_id)
            event_tags = event.tags or []
            await state.update_data(selected_tags=event_tags)
            await callback.message.answer(
                "Выбор тегов:",
                reply_markup=get_admin_tags_selection_keyboard(config.DEFAULT_TAGS, event_tags)
            )
    elif section == "image":
        await callback.message.answer(
            f"Отправьте новое {title}:",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_edit_event_{event_id}")
        )
    else:
        await callback.message.answer(
            f"Введите новое {title}:",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_edit_event_{event_id}")
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_evtag_"), EditEventForm.value)
async def process_edit_event_tags_toggle(callback: CallbackQuery, state: FSMContext):
    action = callback.data.split("_")[2]
    state_data = await state.get_data()
    event_id = state_data["event_id"]
    
    if action == "confirm":
        selected_tags = state_data.get("selected_tags", [])
        async with async_session() as session:
            event = await session.get(Event, event_id)
            if event:
                event.tags = selected_tags
                session.add(event)
                await session.commit()
                
        await state.clear()
        await callback.message.answer(
            "Мероприятие отредактировано!",
            reply_markup=get_after_edit_event_keyboard(event_id)
        )
        await callback.answer()
        return
        
    try:
        idx = int(action)
        tag_name = config.DEFAULT_TAGS[idx]
    except (ValueError, IndexError):
        tag_name = action

    selected = list(state_data.get("selected_tags", []))
    if tag_name in selected:
        selected.remove(tag_name)
    else:
        selected.append(tag_name)
        
    await state.update_data(selected_tags=selected)
    await callback.message.edit_reply_markup(
        reply_markup=get_admin_tags_selection_keyboard(config.DEFAULT_TAGS, selected)
    )
    await callback.answer()


@router.message(EditEventForm.value)
async def process_edit_event_value(message: Message, state: FSMContext):
    state_data = await state.get_data()
    event_id = state_data["event_id"]
    section = state_data["section"]
    
    val = message.html_text if section == "desc" else (message.text.strip() if message.text else "")
    if message.document:
        # Проверка отправки картинки файлом
        mime = message.document.mime_type or ""
        if not mime.startswith("image/"):
            await message.answer("Ошибка: отправленный файл не является изображением. Попробуйте еще раз:")
            return
        
        # Конвертируем документ в фото на стороне ТГ
        import os
        from aiogram.types import FSInputFile
        doc_file_id = message.document.file_id
        file_info = await message.bot.get_file(doc_file_id)
        temp_path = f"temp_{doc_file_id}.jpg"
        await message.bot.download_file(file_info.file_path, temp_path)
        
        sent_msg = await message.answer_photo(photo=FSInputFile(temp_path), caption="Загрузка изображения...")
        photo_file_id = sent_msg.photo[-1].file_id
        await sent_msg.delete()
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
        val = [photo_file_id]
        
    # Валидация формата для дат и времени
    if section == "date" or section == "hidedate":
        try:
            datetime.strptime(val.strip(), "%d.%m.%Y")
        except ValueError:
            await message.answer("Неверный формат даты, попробуйте еще раз (ДД.ММ.ГГГГ):")
            return
    elif section == "time":
        normalized_t = val.strip().replace("—", "-").replace(" - ", "-").replace(" -", "-").replace("- ", "-")
        parts = normalized_t.split("-")
        valid = False
        if len(parts) == 2:
            try:
                datetime.strptime(parts[0].strip(), "%H:%M")
                datetime.strptime(parts[1].strip(), "%H:%M")
                valid = True
            except ValueError:
                pass
        if not valid:
            await message.answer("Неверный формат времени, попробуйте еще раз (ЧЧ:ММ-ЧЧ:ММ):")
            return
    elif section == "hidetime":
        try:
            datetime.strptime(val.strip(), "%H:%M")
        except ValueError:
            await message.answer("Неверный формат времени, попробуйте еще раз (ЧЧ:ММ):")
            return

    send_stream_notif = False
    async with async_session() as session:
        event = await session.get(Event, event_id)
        if event:
            if section == "title":
                event.title = val
            elif section == "titleurl":
                event.title_url = val
            elif section == "date":
                event.date = val
            elif section == "time":
                event.time = val
            elif section == "address":
                event.address = val
            elif section == "desc":
                event.description = val
            elif section == "regurl":
                event.reg_url = val
            elif section == "streamurl":
                had_no_stream = not event.stream_url
                event.stream_url = val
                if had_no_stream and val:
                    send_stream_notif = True
            elif section == "image":
                event.images = val
            elif section == "hidedate":
                event.hide_date = val
            elif section == "hidetime":
                event.hide_time = val
                
            session.add(event)
            await session.commit()
            
            if send_stream_notif:
                asyncio.create_task(send_stream_url_notifications(message.bot, event_id))
            
    await state.clear()
    
    # Показываем карточку после редактирования
    async with async_session() as session:
        event = await session.get(Event, event_id)
        title_html = f"<b>{event.title}</b>"
        if event.title_url:
            title_html = f'<b><a href="{event.title_url}">{event.title}</a></b>'
        tags_str = " ".join([f"#{t}" for t in event.tags])
        desc = event.description or ""
        
        links_list = []
        if event.reg_url:
            links_list.append(f'→ <a href="{event.reg_url}">Регистрация на сайте</a>\nИли в телеграм-боте 👇')
        if event.stream_url:
            links_list.append(f'→ <a href="{event.stream_url}">Трансляция</a>')
        links_str = "\n\n" + "\n".join(links_list) if links_list else ""
        
        text = (
            "Мероприятие отредактировано!\n\n"
            f"{title_html}\n\n"
            f"{desc}\n\n"
            f"📆 <b>Дата:</b> {config.format_display_date(event.date)}\n"
            f"⏳ <b>Время:</b> {event.time}\n"
            f"📍 <b>Место:</b> {event.address}"
            f"{links_str}\n\n"
            f"{tags_str}"
        )
        
        kb = get_after_edit_event_keyboard(event_id)
        if event.images and len(event.images) > 0:
            await message.answer_photo(photo=event.images[0], caption=text, reply_markup=kb, parse_mode="HTML")
        else:
            await message.answer(text, reply_markup=kb, parse_mode="HTML", disable_web_page_preview=True)


# ----- Удаление мероприятия -----

@router.callback_query(F.data == "admin_del_event_list")
async def process_admin_del_event_list(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        active_events = [e for e in events if not is_event_hidden(e)]
        
        await callback.message.answer(
            "Какое мероприятие удалить?",
            reply_markup=get_event_delete_keyboard(active_events)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_confirm_del_event_"))
async def process_admin_confirm_del_event(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[4])
    await callback.message.answer(
        "Вы уверены? Это действие нельзя отменить!",
        reply_markup=get_confirm_delete_event_keyboard(event_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_delete_event_final_"))
async def process_admin_delete_event_final(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[4])
    async with async_session() as session:
        event = await session.get(Event, event_id)
        if event:
            await session.delete(event)
            await session.commit()
            
            await callback.message.answer(
                "Мероприятие удалено!",
                reply_markup=get_after_deleted_event_keyboard()
            )
        else:
            await callback.answer("Мероприятие не найдено.")
    await callback.answer()


# ==========================================
# 3. CRUD ТЕГОВ
# ==========================================

@router.callback_query(F.data == "admin_edit_tags")
async def process_admin_edit_tags(callback: CallbackQuery):
    await callback.message.answer(
        "Редактируем теги. Выберите действие.",
        reply_markup=get_admin_tags_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data == "admin_add_tag")
async def process_admin_add_tag(callback: CallbackQuery, state: FSMContext):
    await state.set_state(TagForm.name)
    await callback.message.answer(
        "Введите название тега без знака “#”.",
        reply_markup=get_cancel_keyboard()
    )
    await callback.answer()


@router.message(TagForm.name)
async def process_save_tag(message: Message, state: FSMContext):
    new_tag = message.text.strip().lstrip('#')
    await state.clear()
    
    from database.models import SystemTag
    async with async_session() as session:
        result = await session.execute(select(SystemTag).where(SystemTag.name == new_tag))
        existing = result.scalar_one_or_none()
        if not existing:
            session.add(SystemTag(name=new_tag))
            await session.commit()
            
    if new_tag not in config.DEFAULT_TAGS:
        config.DEFAULT_TAGS.append(new_tag)
        
    await message.answer(
        "Тег сохранен, спасибо!",
        reply_markup=get_after_tag_created_keyboard()
    )


@router.callback_query(F.data == "admin_del_tag_list")
async def process_admin_del_tag_list(callback: CallbackQuery):
    await callback.message.answer(
        "Какой тег удалить?",
        reply_markup=get_tag_delete_keyboard(config.DEFAULT_TAGS)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_confirm_del_tag_"))
async def process_admin_confirm_del_tag(callback: CallbackQuery):
    idx_str = callback.data.split("_")[4]
    try:
        idx = int(idx_str)
        tag = config.DEFAULT_TAGS[idx]
    except (ValueError, IndexError):
        tag = idx_str

    if tag in config.DEFAULT_TAGS:
        config.DEFAULT_TAGS.remove(tag)
        
    from database.models import SystemTag
    async with async_session() as session:
        # Удаляем тег из базы данных system_tags
        tag_result = await session.execute(select(SystemTag).where(SystemTag.name == tag))
        db_tag = tag_result.scalar_one_or_none()
        if db_tag:
            await session.delete(db_tag)

        users_query = select(User)
        result = await session.execute(users_query)
        for user in result.scalars().all():
            prefs = dict(user.tags_preferences or {})
            if tag in prefs:
                del prefs[tag]
                user.tags_preferences = prefs
                session.add(user)
                
        events_query = select(Event)
        ev_result = await session.execute(events_query)
        for ev in ev_result.scalars().all():
            ev_tags = list(ev.tags or [])
            if tag in ev_tags:
                ev_tags.remove(tag)
                ev.tags = ev_tags
                session.add(ev)
        await session.commit()

    await callback.message.answer(
        "Тег удален!",
        reply_markup=get_after_tag_deleted_keyboard()
    )
    await callback.answer()


# ==========================================
# 4. CRUD РОЗЫГРЫШЕЙ
# ==========================================

@router.callback_query(F.data == "admin_edit_raffles")
async def process_admin_edit_raffles(callback: CallbackQuery):
    await callback.message.answer(
        "Редактируем розыгрыши. Выберите действие.",
        reply_markup=get_admin_raffles_keyboard()
    )
    await callback.answer()


# ----- Добавление розыгрыша -----

@router.callback_query(F.data == "admin_add_raffle")
async def process_admin_add_raffle(callback: CallbackQuery, state: FSMContext):
    await state.set_state(RaffleForm.title)
    await callback.message.answer(
        "Вопрос 1: Название",
        reply_markup=get_skip_keyboard("skip_raf_title")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_raf_title", RaffleForm.title)
async def process_skip_raf_title(callback: CallbackQuery, state: FSMContext):
    await state.update_data(title="")
    await state.set_state(RaffleForm.url)
    await callback.message.answer(
        "Вопрос 2: Ссылка, вшиваемая в название",
        reply_markup=get_skip_keyboard("skip_raf_url", show_back=True, back_callback="back_raf_url")
    )
    await callback.answer()


@router.message(RaffleForm.title)
async def process_add_raffle_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await state.set_state(RaffleForm.url)
    await message.answer(
        "Вопрос 2: Ссылка, вшиваемая в название",
        reply_markup=get_skip_keyboard("skip_raf_url", show_back=True, back_callback="back_raf_url")
    )


@router.callback_query(F.data == "back_raf_url", RaffleForm.url)
async def process_back_raf_url(callback: CallbackQuery, state: FSMContext):
    await state.set_state(RaffleForm.title)
    await callback.message.answer(
        "Вопрос 1: Название",
        reply_markup=get_skip_keyboard("skip_raf_title")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_raf_url", RaffleForm.url)
async def process_skip_raf_url(callback: CallbackQuery, state: FSMContext):
    await state.update_data(url=None)
    await state.set_state(RaffleForm.description)
    await callback.message.answer(
        "Вопрос 3: Основной текст",
        reply_markup=get_skip_keyboard("skip_raf_desc", show_back=True, back_callback="back_raf_desc")
    )
    await callback.answer()


@router.message(RaffleForm.url)
async def process_add_raffle_url(message: Message, state: FSMContext):
    await state.update_data(url=message.text)
    await state.set_state(RaffleForm.description)
    await message.answer(
        "Вопрос 3: Основной текст",
        reply_markup=get_skip_keyboard("skip_raf_desc", show_back=True, back_callback="back_raf_desc")
    )


@router.callback_query(F.data == "back_raf_desc", RaffleForm.description)
async def process_back_raf_desc(callback: CallbackQuery, state: FSMContext):
    await state.set_state(RaffleForm.url)
    await callback.message.answer(
        "Вопрос 2: Ссылка, вшиваемая в название",
        reply_markup=get_skip_keyboard("skip_raf_url", show_back=True, back_callback="back_raf_url")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_raf_desc", RaffleForm.description)
async def process_skip_raf_desc(callback: CallbackQuery, state: FSMContext):
    await state.update_data(description=None)
    await state.set_state(RaffleForm.hide_date)
    await callback.message.answer(
        "Вопрос 4: Дата, когда розыгрыш должен исчезнуть из видимости (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_skip_keyboard("skip_raf_hidedate", show_back=True, back_callback="back_raf_hidedate")
    )
    await callback.answer()


@router.message(RaffleForm.description)
async def process_add_raffle_desc(message: Message, state: FSMContext):
    await state.update_data(description=message.html_text)
    await state.set_state(RaffleForm.hide_date)
    await message.answer(
        "Вопрос 4: Дата, когда розыгрыш должен исчезнуть из видимости (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_skip_keyboard("skip_raf_hidedate", show_back=True, back_callback="back_raf_hidedate")
    )


@router.callback_query(F.data == "back_raf_hidedate", RaffleForm.hide_date)
async def process_back_raf_hidedate(callback: CallbackQuery, state: FSMContext):
    await state.set_state(RaffleForm.description)
    await callback.message.answer(
        "Вопрос 3: Основной текст",
        reply_markup=get_skip_keyboard("skip_raf_desc", show_back=True, back_callback="back_raf_desc")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_raf_hidedate", RaffleForm.hide_date)
async def process_skip_raf_hidedate(callback: CallbackQuery, state: FSMContext):
    await state.update_data(hide_date=None)
    await state.set_state(RaffleForm.hide_time)
    await callback.message.answer(
        "Вопрос 5: Время, когда розыгрыш должен исчезнуть из видимости (строго в формате ЧЧ:ММ, формат 24 часа)",
        reply_markup=get_skip_keyboard("skip_raf_hidetime", show_back=True, back_callback="back_raf_hidetime")
    )
    await callback.answer()


@router.message(RaffleForm.hide_date)
async def process_add_raffle_hide_date(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text.strip(), "%d.%m.%Y")
    except ValueError:
        await message.answer(
            "Неверный формат даты, попробуйте еще раз",
            reply_markup=get_skip_keyboard("skip_raf_hidedate", show_back=True, back_callback="back_raf_hidedate")
        )
        return

    await state.update_data(hide_date=message.text.strip())
    await state.set_state(RaffleForm.hide_time)
    await message.answer(
        "Вопрос 5: Время, когда розыгрыш должен исчезнуть из видимости (строго в формате ЧЧ:ММ, формат 24 часа)",
        reply_markup=get_skip_keyboard("skip_raf_hidetime", show_back=True, back_callback="back_raf_hidetime")
    )


@router.callback_query(F.data == "back_raf_hidetime", RaffleForm.hide_time)
async def process_back_raf_hidetime(callback: CallbackQuery, state: FSMContext):
    await state.set_state(RaffleForm.hide_date)
    await callback.message.answer(
        "Вопрос 4: Дата, когда розыгрыш должен исчезнуть из видимости (строго в формате ЧЧ.ММ.ГГГГ)",
        reply_markup=get_skip_keyboard("skip_raf_hidedate", show_back=True, back_callback="back_raf_hidedate")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_raf_hidetime", RaffleForm.hide_time)
async def process_skip_raf_hidetime(callback: CallbackQuery, state: FSMContext):
    await state.update_data(hide_time=None)
    await ask_raffle_link_decision(callback.message, state)
    await callback.answer()


@router.message(RaffleForm.hide_time)
async def process_add_raffle_hide_time(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text.strip(), "%H:%M")
    except ValueError:
        await message.answer(
            "Неверный формат времени, попробуйте еще раз",
            reply_markup=get_skip_keyboard("skip_raf_hidetime", show_back=True, back_callback="back_raf_hidetime")
        )
        return

    await state.update_data(hide_time=message.text.strip())
    await ask_raffle_link_decision(message, state)


async def ask_raffle_link_decision(message: Message, state: FSMContext):
    await state.set_state(RaffleForm.event_id)
    await message.answer(
        "Хотите привязать розыгрыш к мероприятию? В этом случае он будет виден только тем, "
        "кому это мероприятие подходит по предпочтениям. Если не привязывать, он будет виден абсолютно всем.",
        reply_markup=get_raffle_link_choice_keyboard()
    )


@router.callback_query(F.data == "admin_raffle_link_select", RaffleForm.event_id)
async def process_raffle_link_select(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        
        await callback.message.answer(
            "К какому мероприятию привязать розыгрыш?",
            reply_markup=get_raffle_link_event_keyboard(events)
        )
    await callback.answer()


@router.callback_query(F.data == "admin_raffle_link_back", RaffleForm.event_id)
async def process_raffle_link_back(callback: CallbackQuery, state: FSMContext):
    await ask_raffle_link_decision(callback.message, state)
    await callback.answer()


@router.callback_query(F.data == "admin_raffle_link_none", RaffleForm.event_id)
async def process_save_raffle_no_link(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    
    async with async_session() as session:
        await session.execute(Raffle.__table__.update().values(is_active=0))
        
        new_raffle = Raffle(
            title=data.get("title", ""),
            description=data.get("description"),
            url=data.get("url"),
            hide_date=data.get("hide_date"),
            hide_time=data.get("hide_time"),
            is_active=1,
            event_id=None
        )
        session.add(new_raffle)
        await session.commit()
        
        if callback.message.bot:
            asyncio.create_task(send_raffle_creation_notifications(callback.message.bot, new_raffle))
        
        title_html = new_raffle.title
        if new_raffle.url:
            title_html = f'<a href="{new_raffle.url}">{new_raffle.title}</a>'
        desc = new_raffle.description or ""
        
        await callback.message.answer(
            f"Розыгрыш добавлен!\n\n<b>{title_html}</b>\n\n{desc}",
            reply_markup=get_after_raffle_save_keyboard(),
            parse_mode="HTML",
            disable_web_page_preview=True
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_raffle_link_event_"), RaffleForm.event_id)
async def process_save_raffle_with_link(callback: CallbackQuery, state: FSMContext):
    event_id = int(callback.data.split("_")[4])
    data = await state.get_data()
    await state.clear()
    
    async with async_session() as session:
        await session.execute(Raffle.__table__.update().values(is_active=0))
        
        new_raffle = Raffle(
            title=data.get("title", ""),
            description=data.get("description"),
            url=data.get("url"),
            hide_date=data.get("hide_date"),
            hide_time=data.get("hide_time"),
            is_active=1,
            event_id=event_id
        )
        session.add(new_raffle)
        await session.commit()
        
        if callback.message.bot:
            asyncio.create_task(send_raffle_creation_notifications(callback.message.bot, new_raffle))
        
        event = await session.get(Event, event_id)
        ev_title = event.title if event else ""
        
        title_html = new_raffle.title
        if new_raffle.url:
            title_html = f'<a href="{new_raffle.url}">{new_raffle.title}</a>'
        desc = new_raffle.description or ""
        
        await callback.message.answer(
            f"Розыгрыш добавлен и привязан к мероприятию: {ev_title}\n\n<b>{title_html}</b>\n\n{desc}",
            reply_markup=get_after_raffle_save_keyboard(),
            parse_mode="HTML",
            disable_web_page_preview=True
        )
    await callback.answer()


# ----- Редактирование розыгрыша -----

@router.callback_query(F.data == "admin_edit_raffle_list")
async def process_admin_edit_raffle_list(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Raffle)
        result = await session.execute(query)
        raffles = result.scalars().all()
        
        await callback.message.answer(
            "Какой розыгрыш редактировать?",
            reply_markup=get_raffle_select_edit_keyboard(raffles)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_select_edit_raffle_"))
async def process_admin_select_edit_raffle(callback: CallbackQuery):
    raffle_id = int(callback.data.split("_")[4])
    await callback.message.answer(
        "Какой раздел редактировать?",
        reply_markup=get_raffle_sections_keyboard(raffle_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("edit_rafsec_"))
async def process_edit_rafsec_start(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    raffle_id = int(parts[2])
    section = parts[3]
    
    if section == "eventid":
        async with async_session() as session:
            query = select(Event)
            result = await session.execute(query)
            events = result.scalars().all()
            
            await state.set_state(EditRaffleForm.value)
            await state.update_data(raffle_id=raffle_id, section=section)
            await callback.message.answer(
                "К какому мероприятию привязать розыгрыш?",
                reply_markup=get_raffle_edit_link_event_keyboard(events, raffle_id)
            )
            await callback.answer()
            return
            
    await state.set_state(EditRaffleForm.value)
    await state.update_data(raffle_id=raffle_id, section=section)
    
    section_titles = {
        "title": "Название",
        "url": "Ссылка, вшиваемая в название",
        "desc": "Основной текст",
        "hidedate": "Дата скрытия (ДД.ММ.ГГГГ)",
        "hidetime": "Время скрытия (ЧЧ:ММ)"
    }
    
    title = section_titles.get(section, "значение")
    await callback.message.answer(
        f"Введите новое {title}:",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_edit_raffle_{raffle_id}")
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_raf_edit_link_"))
async def process_edit_raffle_link(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    raffle_id = int(parts[4])
    event_val = parts[5]
    
    await state.clear()
    
    async with async_session() as session:
        raffle = await session.get(Raffle, raffle_id)
        if raffle:
            if event_val == "none":
                raffle.event_id = None
                ev_title = "Не привязано"
            else:
                event_id = int(event_val)
                raffle.event_id = event_id
                event = await session.get(Event, event_id)
                ev_title = event.title if event else "Неизвестное событие"
                
            session.add(raffle)
            await session.commit()
            
            title_html = raffle.title
            if raffle.url:
                title_html = f'<a href="{raffle.url}">{raffle.title}</a>'
            desc = raffle.description or ""
            
            await callback.message.answer(
                f"Розыгрыш отредактирован и привязан к мероприятию: {ev_title}\n\n<b>{title_html}</b>\n\n{desc}",
                reply_markup=get_after_edit_raffle_keyboard(),
                parse_mode="HTML",
                disable_web_page_preview=True
            )
        else:
            await callback.answer("Розыгрыш не найден.")
    await callback.answer()


@router.message(EditRaffleForm.value)
async def process_edit_raffle_value(message: Message, state: FSMContext):
    state_data = await state.get_data()
    raffle_id = state_data["raffle_id"]
    section = state_data["section"]
    val = message.html_text if section == "desc" else message.text.strip()
    
    if section == "hidedate":
        try:
            datetime.strptime(val, "%d.%m.%Y")
        except ValueError:
            await message.answer("Неверный формат даты, попробуйте еще раз (ДД.ММ.ГГГГ):")
            return
    elif section == "hidetime":
        try:
            datetime.strptime(val, "%H:%M")
        except ValueError:
            await message.answer("Неверный формат времени, попробуйте еще раз (ЧЧ:ММ):")
            return

    async with async_session() as session:
        raffle = await session.get(Raffle, raffle_id)
        if raffle:
            if section == "title":
                raffle.title = val
            elif section == "url":
                raffle.url = val
            elif section == "desc":
                raffle.description = val
            elif section == "hidedate":
                raffle.hide_date = val
            elif section == "hidetime":
                raffle.hide_time = val
            session.add(raffle)
            await session.commit()
            
    await state.clear()
    
    async with async_session() as session:
        raffle = await session.get(Raffle, raffle_id)
        title_html = raffle.title
        if raffle.url:
            title_html = f'<a href="{raffle.url}">{raffle.title}</a>'
        desc = raffle.description or ""
        
        await message.answer(
            f"Розыгрыш отредактирован!\n\n<b>{title_html}</b>\n\n{desc}",
            reply_markup=get_after_edit_raffle_keyboard(),
            parse_mode="HTML",
            disable_web_page_preview=True
        )


# ----- Удаление розыгрыша -----

@router.callback_query(F.data == "admin_del_raffle_list")
async def process_admin_del_raffle_list(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Raffle)
        result = await session.execute(query)
        raffles = result.scalars().all()
        
        await callback.message.answer(
            "Какой розыгрыш удалить?",
            reply_markup=get_raffle_delete_keyboard(raffles)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_confirm_del_raffle_"))
async def process_admin_confirm_del_raffle(callback: CallbackQuery):
    raffle_id = int(callback.data.split("_")[4])
    await callback.message.answer(
        "Вы уверены? Это действие нельзя отменить!",
        reply_markup=get_confirm_delete_raffle_keyboard(raffle_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_delete_raffle_final_"))
async def process_admin_delete_raffle_final(callback: CallbackQuery):
    raffle_id = int(callback.data.split("_")[4])
    async with async_session() as session:
        raffle = await session.get(Raffle, raffle_id)
        if raffle:
            await session.delete(raffle)
            await session.commit()
            
            await callback.message.answer(
                "Розыгрыш удален!",
                reply_markup=get_after_deleted_raffle_keyboard()
            )
        else:
            await callback.answer("Розыгрыш не найден.")
    await callback.answer()


# ==========================================
# 5. ПРАВА АДМИНИСТРАТОРА
# ==========================================

@router.callback_query(F.data == "admin_rights")
async def process_admin_rights(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Admin)
        result = await session.execute(query)
        admins = result.scalars().all()
        
        admin_list = [f"@{config.SUPER_ADMIN_USERNAME}"]
        for a in admins:
            if a.username.lower() != config.SUPER_ADMIN_USERNAME.lower():
                admin_list.append(f"@{a.username}")
                
        admin_text = "\n".join(admin_list)
        text = f"Администраторы бота:\n\n{admin_text}"
        
        await callback.message.answer(
            text,
            reply_markup=get_admin_rights_keyboard()
        )
    await callback.answer()


@router.callback_query(F.data == "admin_del_rights_list")
async def process_admin_del_rights_list(callback: CallbackQuery):
    async with async_session() as session:
        query = select(Admin)
        result = await session.execute(query)
        admins = result.scalars().all()
        
        await callback.message.answer(
            "Выберите пользователя, которого хотите лишить статуса администратора.",
            reply_markup=get_admin_delete_select_keyboard(admins)
        )
    await callback.answer()


@router.callback_query(F.data == "admin_add_rights")
async def process_admin_add_rights_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminRightsForm.username)
    await callback.message.answer(
        "Введите Telegram-никнейм нового администратора (без символа @):",
        reply_markup=get_cancel_keyboard()
    )
    await callback.answer()


@router.message(AdminRightsForm.username)
async def process_save_admin_rights(message: Message, state: FSMContext):
    username = message.text.strip().lstrip('@')
    await state.clear()
    
    if username.lower() == config.SUPER_ADMIN_USERNAME.lower():
        await message.answer(
            "Этот пользователь является суперадминистратором по умолчанию.",
            reply_markup=get_admin_main_keyboard()
        )
        return
        
    async with async_session() as session:
        q = select(Admin).where(func.lower(Admin.username) == username.lower())
        res = await session.execute(q)
        existing = res.scalar_one_or_none()
        
        if existing:
            await message.answer(
                f"Пользователь @{username} уже является администратором.",
                reply_markup=get_admin_main_keyboard()
            )
            return
            
        new_admin = Admin(username=username)
        session.add(new_admin)
        await session.commit()
        
    await message.answer(
        f"Пользователь @{username} успешно добавлен в список администраторов.",
        reply_markup=get_admin_main_keyboard()
    )


@router.callback_query(F.data.startswith("admin_del_rights_"))
async def process_admin_del_rights(callback: CallbackQuery):
    admin_id = int(callback.data.split("_")[3])
    
    async with async_session() as session:
        admin = await session.get(Admin, admin_id)
        if admin:
            username = admin.username
            if username.lower() == config.SUPER_ADMIN_USERNAME.lower():
                await callback.answer("Ошибка: Нельзя удалить суперадминистратора!", show_alert=True)
                return
                
            await session.delete(admin)
            await session.commit()
            await callback.message.answer(
                f"Права администратора для @{username} аннулированы.",
                reply_markup=get_admin_main_keyboard()
            )
        else:
            await callback.answer("Администратор не найден.")
    await callback.answer()


# ==========================================
# 6. CRUD ПОСТ-МАТЕРИАЛОВ
# ==========================================

@router.callback_query(F.data == "admin_edit_post_mats")
async def process_admin_edit_post_mats(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer(
        "Редактируем пост-материалы. Выберите действие.",
        reply_markup=get_admin_post_mats_keyboard()
    )
    await callback.answer()


# ----- Добавление пост-материалов -----

@router.callback_query(F.data == "admin_add_post_mats")
async def process_admin_add_post_mats(callback: CallbackQuery):
    await callback.message.answer(
        "К какому мероприятию добавить пост-материалы? Выберите тематику.",
        reply_markup=get_post_mats_tags_keyboard(config.DEFAULT_TAGS, is_delete=False)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("pm_tag_add_"))
async def process_pm_tag_add_select(callback: CallbackQuery):
    idx_str = callback.data.split("_")[3]
    try:
        idx = int(idx_str)
        tag = config.DEFAULT_TAGS[idx]
    except (ValueError, IndexError):
        tag = idx_str
        
    async with async_session() as session:
        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        
        tagged_events = [e for e in events if tag in (e.tags or [])]
        
        if not tagged_events:
            await callback.message.answer(
                f"Мероприятий с тегом #{tag} не найдено.",
                reply_markup=get_admin_post_mats_keyboard()
            )
        else:
            await callback.message.answer(
                "Выберите мероприятие для добавления пост-материалов:",
                reply_markup=get_admin_post_mats_event_select_keyboard(tagged_events, is_delete=False)
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_pm_add_"))
async def process_admin_pm_add_start(callback: CallbackQuery, state: FSMContext):
    event_id = int(callback.data.split("_")[3])
    await state.set_state(PostMatsForm.photos_url)
    await state.update_data(event_id=event_id)
    
    await callback.message.answer(
        "Вопрос 1: Вставьте ссылку на фотографии с мероприятия",
        reply_markup=get_skip_keyboard("skip_pm_photos")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_pm_photos", PostMatsForm.photos_url)
async def process_skip_pm_photos(callback: CallbackQuery, state: FSMContext):
    await state.update_data(photos_url=None)
    await state.set_state(PostMatsForm.photographer_name)
    await callback.message.answer(
        "Вопрос 2: Укажите имя фотографа",
        reply_markup=get_skip_keyboard("skip_pm_photographer_name", show_back=True, back_callback="back_pm_photographer_name")
    )
    await callback.answer()


@router.message(PostMatsForm.photos_url)
async def process_add_pm_photos(message: Message, state: FSMContext):
    await state.update_data(photos_url=message.text.strip())
    await state.set_state(PostMatsForm.photographer_name)
    await message.answer(
        "Вопрос 2: Укажите имя фотографа",
        reply_markup=get_skip_keyboard("skip_pm_photographer_name", show_back=True, back_callback="back_pm_photographer_name")
    )


@router.callback_query(F.data == "back_pm_photographer_name", PostMatsForm.photographer_name)
async def process_back_pm_photographer_name(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PostMatsForm.photos_url)
    await callback.message.answer(
        "Вопрос 1: Вставьте ссылку на фотографии с мероприятия",
        reply_markup=get_skip_keyboard("skip_pm_photos")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_pm_photographer_name", PostMatsForm.photographer_name)
async def process_skip_pm_photographer_name(callback: CallbackQuery, state: FSMContext):
    await state.update_data(photographer_name=None, photographer_url=None)
    await state.set_state(PostMatsForm.stream_record_url)
    await callback.message.answer(
        "Вопрос 2: Вставьте ссылку на запись трансляции",
        reply_markup=get_skip_keyboard("skip_pm_stream", show_back=True, back_callback="back_pm_stream")
    )
    await callback.answer()


@router.message(PostMatsForm.photographer_name)
async def process_add_pm_photographer_name(message: Message, state: FSMContext):
    await state.update_data(photographer_name=message.text.strip())
    await state.set_state(PostMatsForm.photographer_url)
    await message.answer(
        "Вопрос 3: Укажите ссылку, вшиваемую в имя фотографа",
        reply_markup=get_skip_keyboard("skip_pm_photographer_url", show_back=True, back_callback="back_pm_photographer_url")
    )


@router.callback_query(F.data == "back_pm_photographer_url", PostMatsForm.photographer_url)
async def process_back_pm_photographer_url(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PostMatsForm.photographer_name)
    await callback.message.answer(
        "Вопрос 2: Укажите имя фотографа",
        reply_markup=get_skip_keyboard("skip_pm_photographer_name", show_back=True, back_callback="back_pm_photographer_name")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_pm_photographer_url", PostMatsForm.photographer_url)
async def process_skip_pm_photographer_url(callback: CallbackQuery, state: FSMContext):
    await state.update_data(photographer_url=None)
    await state.set_state(PostMatsForm.stream_record_url)
    await callback.message.answer(
        "Вопрос 4: Вставьте ссылку на запись трансляции",
        reply_markup=get_skip_keyboard("skip_pm_stream", show_back=True, back_callback="back_pm_stream")
    )
    await callback.answer()


@router.message(PostMatsForm.photographer_url)
async def process_add_pm_photographer_url(message: Message, state: FSMContext):
    link = message.text.strip()
    if not (link.startswith("http://") or link.startswith("https://")):
        await message.answer(
            "Некорректная ссылка, введите еще раз (должна начинаться с http:// или https://):",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_pm_photographer_url")
        )
        return
    await state.update_data(photographer_url=link)
    await state.set_state(PostMatsForm.stream_record_url)
    await message.answer(
        "Вопрос 4: Вставьте ссылку на запись трансляции",
        reply_markup=get_skip_keyboard("skip_pm_stream", show_back=True, back_callback="back_pm_stream")
    )


@router.callback_query(F.data == "back_pm_stream", PostMatsForm.stream_record_url)
async def process_back_pm_stream(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if data.get("photographer_name") is None:
        await state.set_state(PostMatsForm.photographer_name)
        await callback.message.answer(
            "Вопрос 2: Укажите имя фотографа",
            reply_markup=get_skip_keyboard("skip_pm_photographer_name", show_back=True, back_callback="back_pm_photographer_name")
        )
    else:
        await state.set_state(PostMatsForm.photographer_url)
        await callback.message.answer(
            "Вопрос 3: Укажите ссылку, вшиваемую в имя фотографа",
            reply_markup=get_skip_keyboard("skip_pm_photographer_url", show_back=True, back_callback="back_pm_photographer_url")
        )
    await callback.answer()


@router.callback_query(F.data == "skip_pm_stream", PostMatsForm.stream_record_url)
async def process_skip_pm_stream(callback: CallbackQuery, state: FSMContext):
    await state.update_data(stream_record_url=None)
    await state.set_state(PostMatsForm.article_url)
    data = await state.get_data()
    q_num = 5 if data.get("photographer_name") is not None else 3
    await callback.message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на статью-конспект",
        reply_markup=get_skip_keyboard("skip_pm_article", show_back=True, back_callback="back_pm_article")
    )
    await callback.answer()


@router.message(PostMatsForm.stream_record_url)
async def process_add_pm_stream(message: Message, state: FSMContext):
    await state.update_data(stream_record_url=message.text.strip())
    await state.set_state(PostMatsForm.article_url)
    data = await state.get_data()
    q_num = 5 if data.get("photographer_name") is not None else 3
    await message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на статью-конспект",
        reply_markup=get_skip_keyboard("skip_pm_article", show_back=True, back_callback="back_pm_article")
    )


@router.callback_query(F.data == "back_pm_article", PostMatsForm.article_url)
async def process_back_pm_article(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PostMatsForm.stream_record_url)
    data = await state.get_data()
    q_num = 4 if data.get("photographer_name") is not None else 2
    await callback.message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на запись трансляции",
        reply_markup=get_skip_keyboard("skip_pm_stream", show_back=True, back_callback="back_pm_stream")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_pm_article", PostMatsForm.article_url)
async def process_skip_pm_article(callback: CallbackQuery, state: FSMContext):
    await state.update_data(article_url=None)
    await state.set_state(PostMatsForm.presentations_url)
    data = await state.get_data()
    q_num = 6 if data.get("photographer_name") is not None else 4
    await callback.message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на презентации спикеров",
        reply_markup=get_skip_keyboard("skip_pm_presentations", show_back=True, back_callback="back_pm_presentations")
    )
    await callback.answer()


@router.message(PostMatsForm.article_url)
async def process_add_pm_article(message: Message, state: FSMContext):
    await state.update_data(article_url=message.text.strip())
    await state.set_state(PostMatsForm.presentations_url)
    data = await state.get_data()
    q_num = 6 if data.get("photographer_name") is not None else 4
    await message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на презентации спикеров",
        reply_markup=get_skip_keyboard("skip_pm_presentations", show_back=True, back_callback="back_pm_presentations")
    )


@router.callback_query(F.data == "back_pm_presentations", PostMatsForm.presentations_url)
async def process_back_pm_presentations(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PostMatsForm.article_url)
    data = await state.get_data()
    q_num = 5 if data.get("photographer_name") is not None else 3
    await callback.message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на статью-конспект",
        reply_markup=get_skip_keyboard("skip_pm_article", show_back=True, back_callback="back_pm_article")
    )
    await callback.answer()


@router.callback_query(F.data == "skip_pm_presentations", PostMatsForm.presentations_url)
async def process_skip_pm_presentations(callback: CallbackQuery, state: FSMContext):
    await state.update_data(presentations_url=None)
    await state.set_state(PostMatsForm.other_materials_url)
    data = await state.get_data()
    q_num = 7 if data.get("photographer_name") is not None else 5
    await callback.message.answer(
        f"Вопрос {q_num}: Если хотите, вставьте ссылку на другие материалы или завершите добавление",
        reply_markup=get_skip_keyboard("finish_pm_adding", show_back=True, back_callback="back_pm_other")
    )
    await callback.answer()


@router.message(PostMatsForm.presentations_url)
async def process_add_pm_presentations(message: Message, state: FSMContext):
    await state.update_data(presentations_url=message.text.strip())
    await state.set_state(PostMatsForm.other_materials_url)
    data = await state.get_data()
    q_num = 7 if data.get("photographer_name") is not None else 5
    await message.answer(
        f"Вопрос {q_num}: Если хотите, вставьте ссылку на другие материалы или завершите добавление",
        reply_markup=get_skip_keyboard("finish_pm_adding", show_back=True, back_callback="back_pm_other")
    )


@router.callback_query(F.data == "back_pm_other", PostMatsForm.other_materials_url)
async def process_back_pm_other(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PostMatsForm.presentations_url)
    data = await state.get_data()
    q_num = 6 if data.get("photographer_name") is not None else 4
    await callback.message.answer(
        f"Вопрос {q_num}: Вставьте ссылку на презентации спикеров",
        reply_markup=get_skip_keyboard("skip_pm_presentations", show_back=True, back_callback="back_pm_presentations")
    )
    await callback.answer()


@router.message(PostMatsForm.other_materials_url)
async def process_add_pm_other(message: Message, state: FSMContext):
    await state.update_data(other_materials_url=message.text.strip())
    await save_post_mats_to_db(message, state)


@router.callback_query(F.data == "finish_pm_adding", PostMatsForm.other_materials_url)
async def process_finish_pm_adding_btn(callback: CallbackQuery, state: FSMContext):
    await state.update_data(other_materials_url=None)
    await save_post_mats_to_db(callback.message, state)
    await callback.answer()


async def send_post_mats_notifications(bot, event_id: int):
    from database.db import async_session
    from database.models import User, Registration, Event
    from sqlalchemy import select
    import logging
    logger = logging.getLogger("post_mats_notifications")

    async with async_session() as session:
        event = await session.get(Event, event_id)
        if not event:
            return

        # Проверяем, есть ли действительно добавленные пост-материалы
        if not any([event.photos_url, event.stream_record_url, event.article_url, event.presentations_url, event.other_materials_url]):
            return

        stmt_users = select(User)
        res_users = await session.execute(stmt_users)
        users = res_users.scalars().all()

        stmt_regs = select(Registration.user_id).where(Registration.event_id == event_id)
        res_regs = await session.execute(stmt_regs)
        registered_user_ids = set(res_regs.scalars().all())

        title = event.title
        title_url = event.title_url
        tags = event.tags or []
        images = event.images or []

        links = []
        if event.photos_url:
            if event.photographer_name:
                if event.photographer_url:
                    photo_str = f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a> (Фотограф: <a href="{event.photographer_url}">{event.photographer_name}</a>)'
                else:
                    photo_str = f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a> (Фотограф: {event.photographer_name})'
            else:
                photo_str = f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a>'
            links.append(photo_str)
        if event.stream_record_url:
            links.append(f'▶️ Запись трансляции: <a href="{event.stream_record_url}">открыть</a>')
        if event.article_url:
            links.append(f'✍️ Статья-конспект: <a href="{event.article_url}">открыть</a>')
        if event.presentations_url:
            links.append(f'🖼 Презентации спикеров: <a href="{event.presentations_url}">открыть</a>')
        if event.other_materials_url:
            links.append(f'📎 Другие материалы: <a href="{event.other_materials_url}">открыть</a>')

        mats_str = "\n".join(links)
        tags_str = " ".join([f"#{tag}" for tag in tags])

        title_html = f"<b>{title}</b>"
        if title_url:
            title_html = f'<b><a href="{title_url}">{title}</a></b>'

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    user_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="← К списку", callback_data="back_to_main")]
    ])

    pm_trigger = "После мероприятия, когда появятся пост-материалы"

    for user in users:
        user_notifications = user.notification_preferences or {}
        if user_notifications.get(pm_trigger, True) is False:
            continue

        is_registered = user.telegram_id in registered_user_ids
        user_tags = user.tags_preferences or {}
        has_matching_tag = any(user_tags.get(t) for t in tags)

        if is_registered:
            header_text = "Появились пост-материалы по мероприятию, на которое вы регистрировались!\n\n"
        elif has_matching_tag:
            header_text = "Появились пост-материалы по мероприятию из ваших предпочтений!\n\n"
        else:
            continue

        notification_text = (
            f"{header_text}"
            f"{title_html}\n\n"
            f"{mats_str}\n\n"
            f"{tags_str}"
        )

        try:
            if images and len(images) > 0:
                await bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=images[0],
                    caption=notification_text,
                    parse_mode="HTML",
                    reply_markup=user_kb
                )
            else:
                await bot.send_message(
                    chat_id=user.telegram_id,
                    text=notification_text,
                    parse_mode="HTML",
                    reply_markup=user_kb,
                    disable_web_page_preview=True
                )
            await asyncio.sleep(0.05)
        except Exception as e:
            logger.error(f"Failed to send post-mats notification to user {user.telegram_id}: {e}")


async def save_post_mats_to_db(message: Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    
    event_id = data["event_id"]
    async with async_session() as session:
        event = await session.get(Event, event_id)
        if event:
            event.photos_url = data.get("photos_url")
            event.photographer_name = data.get("photographer_name")
            event.photographer_url = data.get("photographer_url")
            event.stream_record_url = data.get("stream_record_url")
            event.article_url = data.get("article_url")
            event.presentations_url = data.get("presentations_url")
            event.other_materials_url = data.get("other_materials_url")
            session.add(event)
            await session.commit()
            
            links = []
            if event.photos_url:
                if event.photographer_name:
                    if event.photographer_url:
                        photo_str = f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a> (Фотограф: <a href="{event.photographer_url}">{event.photographer_name}</a>)'
                    else:
                        photo_str = f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a> (Фотограф: {event.photographer_name})'
                else:
                    photo_str = f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a>'
                links.append(photo_str)
            if event.stream_record_url:
                links.append(f'▶️ Запись трансляции: <a href="{event.stream_record_url}">открыть</a>')
            if event.article_url:
                links.append(f'✍️ Статья-конспект: <a href="{event.article_url}">открыть</a>')
            if event.presentations_url:
                links.append(f'🖼 Презентации спикеров: <a href="{event.presentations_url}">открыть</a>')
            if event.other_materials_url:
                links.append(f'📎 Другие материалы: <a href="{event.other_materials_url}">открыть</a>')
                
            mats_str = "\n".join(links) if links else "Материалы не добавлены."
            
            bot = getattr(message, 'bot', None)
            if bot and links:
                asyncio.create_task(send_post_mats_notifications(bot, event_id))

            await message.answer(
                f"Пост-материалы добавлены!\n\n{mats_str}",
                reply_markup=get_after_post_mats_saved_keyboard(),
                parse_mode="HTML",
                disable_web_page_preview=True
            )


# ----- Удаление пост-материалов -----

@router.callback_query(F.data == "admin_del_post_mats")
async def process_admin_del_post_mats(callback: CallbackQuery):
    await callback.message.answer(
        "Пост-материалы какого мероприятия удалить? Выберите тематику.",
        reply_markup=get_post_mats_tags_keyboard(config.DEFAULT_TAGS, is_delete=True)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("pm_tag_del_"))
async def process_pm_tag_del_select(callback: CallbackQuery):
    idx_str = callback.data.split("_")[3]
    try:
        idx = int(idx_str)
        tag = config.DEFAULT_TAGS[idx]
    except (ValueError, IndexError):
        tag = idx_str
        
    async with async_session() as session:
        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        
        tagged_events = []
        for e in events:
            if tag in (e.tags or []):
                if any([e.photos_url, e.stream_record_url, e.article_url, e.presentations_url, e.other_materials_url]):
                    tagged_events.append(e)
                    
        if not tagged_events:
            await callback.message.answer(
                f"Мероприятий с пост-материалами по тегу #{tag} не найдено.",
                reply_markup=get_admin_post_mats_keyboard()
            )
        else:
            await callback.message.answer(
                "Выберите мероприятие для удаления его пост-материалов.",
                reply_markup=get_admin_post_mats_event_select_keyboard(tagged_events, is_delete=True)
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_pm_del_confirm_"))
async def process_admin_pm_del_final(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[4])
    
    async with async_session() as session:
        event = await session.get(Event, event_id)
        if event:
            event.photos_url = None
            event.stream_record_url = None
            event.article_url = None
            event.presentations_url = None
            event.other_materials_url = None
            session.add(event)
            await session.commit()
            
            await callback.message.answer(
                "Пост-материалы удалены!",
                reply_markup=get_after_post_mats_saved_keyboard()
            )
        else:
            await callback.answer("Мероприятие не найдено.")
    await callback.answer()


@router.callback_query(F.data.startswith("admin_pm_del_"), ~F.data.startswith("admin_pm_del_confirm_"))
async def process_admin_pm_del_confirm_start(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[3])
    
    async with async_session() as session:
        event = await session.get(Event, event_id)
        if not event:
            await callback.answer("Мероприятие не найдено.")
            return
            
        links = []
        if event.photos_url:
            links.append(f'📸 Фотографии с мероприятия: <a href="{event.photos_url}">открыть</a>')
        if event.stream_record_url:
            links.append(f'▶️ Запись трансляции: <a href="{event.stream_record_url}">открыть</a>')
        if event.article_url:
            links.append(f'✍️ Статья-конспект: <a href="{event.article_url}">открыть</a>')
        if event.presentations_url:
            links.append(f'🖼 Презентации спикеров: <a href="{event.presentations_url}">открыть</a>')
        if event.other_materials_url:
            links.append(f'📎 Другие материалы: <a href="{event.other_materials_url}">открыть</a>')
            
        mats_str = "\n".join(links) if links else "Материалы отсутствуют."
        
        await callback.message.answer(
            f"Вы уверены? Это действие нельзя отменить!\n\n{mats_str}",
            reply_markup=get_confirm_del_post_mats_keyboard(event_id),
            parse_mode="HTML",
            disable_web_page_preview=True
        )
    await callback.answer()



async def send_stream_url_notifications(bot, event_id: int):
    from database.db import async_session
    from database.models import User, Registration, Event
    from sqlalchemy import select
    import logging
    logger = logging.getLogger("stream_notifications")
    
    async with async_session() as session:
        event = await session.get(Event, event_id)
        if not event or not event.stream_url:
            return
            
        stmt = select(Registration).where(Registration.event_id == event.id, Registration.status == "удаленно")
        result = await session.execute(stmt)
        registrations = result.scalars().all()
        
        user_ids = [r.user_id for r in registrations]
        if not user_ids:
            return
            
        stmt_users = select(User).where(User.telegram_id.in_(user_ids))
        result_users = await session.execute(stmt_users)
        users = result_users.scalars().all()
        
        # Capture variables safely inside the session
        title = event.title
        title_url = event.title_url
        date = event.date
        time = event.time
        stream_url = event.stream_url
        tags = event.tags or []
        images = event.images or []
        
    title_html = f"<b>{title}</b>"
    if title_url:
        title_html = f'<b><a href="{title_url}">{title}</a></b>'
        
    tags_str = " ".join([f"#{tag}" for tag in tags])
    
    notification_text = (
        "Ссылка на трансляцию мероприятия уже доступна!\n\n"
        f"Вы зарегистрировались на трансляцию мероприятия {title_html}.\n\n"
        "Высылаем ссылку на подключение:\n\n"
        f"📆 Дата: {config.format_display_date(date)}\n"
        f"⏳ Время: {time}\n\n"
        f'→ <a href="{stream_url}">Трансляция</a>\n\n'
        f"{tags_str}"
    )
    
    for user in users:
        try:
            if images and len(images) > 0:
                await bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=images[0],
                    caption=notification_text,
                    parse_mode="HTML"
                )
            else:
                await bot.send_message(
                    chat_id=user.telegram_id,
                    text=notification_text,
                    parse_mode="HTML",
                    disable_web_page_preview=True
                )
            await asyncio.sleep(0.05)
        except Exception as e:
            logger.error(f"Failed to send stream URL notification to user {user.telegram_id}: {e}")


async def send_event_creation_notifications(bot, event):
    from database.db import async_session
    from database.models import User
    from sqlalchemy import select
    import logging
    logger = logging.getLogger("event_notifications")
    
    title_html = f"<b>{event.title}</b>"
    if event.title_url:
        title_html = f'<b><a href="{event.title_url}">{event.title}</a></b>'
        
    tags_str = " ".join([f"#{tag}" for tag in event.tags])
    links_list = []
    if event.reg_url:
        links_list.append(f'→ <a href="{event.reg_url}">Регистрация на сайте</a>\nИли в телеграм-боте 👇')
    if event.stream_url:
        links_list.append(f'→ <a href="{event.stream_url}">Трансляция</a>')
    links_str = "\n\n" + "\n".join(links_list) if links_list else ""

    notification_text = (
        f"Появилось мероприятие по вашим предпочтениям!\n\n"
        f"{title_html}\n\n"
        f"📆 <b>Дата:</b> {config.format_display_date(event.date)}\n"
        f"⏳ <b>Время:</b> {event.time}\n"
        f"📍 <b>Место:</b> {event.address}"
        f"{links_str}\n\n"
        f"{tags_str}"
    )
    
    async with async_session() as session:
        result = await session.execute(select(User))
        users = result.scalars().all()
        
    for user in users:
        user_prefs = user.notification_preferences or {}
        if not user_prefs.get("Появилось мероприятие по избранной теме", True):
            continue
            
        user_tags = user.tags_preferences or {}
        has_matching_tag = any(user_tags.get(t, False) for t in event.tags)
        if not has_matching_tag:
            continue
            
        try:
            if event.images and len(event.images) > 0:
                await bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=event.images[0],
                    caption=notification_text,
                    reply_markup=get_event_notification_keyboard(event.id),
                    parse_mode="HTML"
                )
            else:
                await bot.send_message(
                    chat_id=user.telegram_id,
                    text=notification_text,
                    reply_markup=get_event_notification_keyboard(event.id),
                    parse_mode="HTML",
                    disable_web_page_preview=True
                )
            await asyncio.sleep(0.05)
        except Exception as e:
            logger.error(f"Failed to send event notification to user {user.telegram_id}: {e}")


async def send_raffle_creation_notifications(bot, raffle):
    from database.db import async_session
    from database.models import User, Event
    from sqlalchemy import select
    import logging
    logger = logging.getLogger("raffle_notifications")

    title_html = raffle.title
    if raffle.url:
        title_html = f'<a href="{raffle.url}">{raffle.title}</a>'

    desc = raffle.description or ""
    notification_text = (
        f"Появился розыгрыш по вашим предпочтениям!\n\n"
        f"<b>{title_html}</b>\n\n"
        f"{desc}"
    )

    async with async_session() as session:
        event = None
        if raffle.event_id is not None:
            event = await session.get(Event, raffle.event_id)
        
        result = await session.execute(select(User))
        users = result.scalars().all()

    for user in users:
        if event is not None:
            user_tags = user.tags_preferences or {}
            has_matching_tag = any(user_tags.get(t, False) for t in (event.tags or []))
            if not has_matching_tag:
                continue

        try:
            await bot.send_message(
                chat_id=user.telegram_id,
                text=notification_text,
                reply_markup=get_to_main_keyboard(),
                parse_mode="HTML",
                disable_web_page_preview=True
            )
            await asyncio.sleep(0.05)
        except Exception as e:
            logger.error(f"Failed to send raffle notification to user {user.telegram_id}: {e}")


# ==========================================
# ВЫГРУЗКА РЕГИСТРАЦИЙ В EXCEL
# ==========================================

@router.callback_query(F.data == "admin_export_registrations")
async def process_admin_export_registrations(callback: CallbackQuery):
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return
        
        res = await session.execute(select(EventSeries))
        series_list = res.scalars().all()

        await callback.message.answer(
            "Выберите категорию мероприятий для выгрузки списка регистраций:",
            reply_markup=get_admin_export_type_keyboard(series_list)
        )
    await callback.answer()


@router.callback_query(F.data == "admin_export_select_active")
async def process_admin_export_select_active(callback: CallbackQuery):
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        active_events = [e for e in events if not is_event_hidden(e)]

        if not active_events:
            await callback.message.answer(
                "Нет активных мероприятий.",
                reply_markup=get_admin_export_type_keyboard()
            )
        else:
            reg_counts = {}
            for e in active_events:
                cnt_query = select(func.count(Registration.user_id)).where(Registration.event_id == e.id)
                cnt_res = await session.execute(cnt_query)
                reg_counts[e.id] = cnt_res.scalar() or 0
                
            await callback.message.answer(
                "Выберите активное мероприятие для выгрузки отчета:",
                reply_markup=get_admin_export_events_keyboard(active_events, reg_counts)
            )
    await callback.answer()


@router.callback_query(F.data == "admin_export_select_archive")
async def process_admin_export_select_archive(callback: CallbackQuery):
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        await callback.message.answer(
            "Выберите тему для поиска архивных мероприятий:",
            reply_markup=get_admin_export_archive_tags_keyboard(config.DEFAULT_TAGS)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_exarchtag_"))
async def process_admin_exarchtag_select(callback: CallbackQuery):
    tag_idx = int(callback.data.split("_")[2])
    try:
        tag = config.DEFAULT_TAGS[tag_idx]
    except IndexError:
        await callback.answer("Тема не найдена.")
        return

    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа.", show_alert=True)
            return

        query = select(Event)
        result = await session.execute(query)
        events = result.scalars().all()
        archived_events = [e for e in events if is_event_hidden(e) and tag in (e.tags or [])]

        if not archived_events:
            await callback.message.answer(
                f"Нет архивных мероприятий по теме «{tag}».",
                reply_markup=get_admin_export_archive_tags_keyboard(config.DEFAULT_TAGS)
            )
        else:
            reg_counts = {}
            for e in archived_events:
                cnt_query = select(func.count(Registration.user_id)).where(Registration.event_id == e.id)
                cnt_res = await session.execute(cnt_query)
                reg_counts[e.id] = cnt_res.scalar() or 0
                
            await callback.message.answer(
                f"Архивные мероприятия по теме «{tag}»:\nВыберите мероприятие для выгрузки отчета.",
                reply_markup=get_admin_export_archive_events_keyboard(archived_events, tag_idx, reg_counts)
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_export_event_"))
async def process_admin_export_event(callback: CallbackQuery):
    import io
    import openpyxl
    
    event_id = int(callback.data.split("_")[3])
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        event = await session.get(Event, event_id)
        if not event:
            await callback.answer("Мероприятие не найдено.")
            return

        # Fetch registrations for this event
        regs_query = select(Registration).where(Registration.event_id == event_id)
        regs_result = await session.execute(regs_query)
        registrations = list(regs_result.scalars().all())
        
        # Sort registrations so that "думаю" status is always at the end
        registrations.sort(key=lambda r: 1 if r.status == "думаю" else 0)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Регистрации"

        # Headers
        headers = [
            "ФИО",
            "Электронная почта",
            "Телефон",
            "Откуда вы узнали о событии?",
            "Вы планируете быть очно или удалённо?",
            "Дата",
            "TG ник"
        ]
        ws.append(headers)

        for reg in registrations:
            user = await session.get(User, reg.user_id)
            fio = getattr(user, "full_name", "-") or "-"
            email = getattr(user, "email", "-") or "-"
            phone = getattr(user, "phone", "-") or "-"
            source = "Регистрация через ТГ-бот Хаба"
            status = reg.status
            reg_date = reg.registration_date or "-"
            
            tg_username = ""
            if user and user.username:
                username = user.username.strip()
                if username:
                    tg_username = f"@{username.lstrip('@')}"
            else:
                tg_username = "-"

            ws.append([fio, email, phone, source, status, reg_date, tg_username])

        # Adjust column widths dynamically
        for col in ws.columns:
            vals = [cell.value for cell in col]
            max_len = max(len(str(val or '')) for val in vals) if vals else 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Save to memory stream
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Send file to admin
        safe_title = "".join(c for c in event.title if c.isalnum() or c in (" ", "_", "-")).strip()
        filename = f"registrations_{event_id}_{safe_title}.xlsx"
        
        input_file = BufferedInputFile(file_stream.read(), filename=filename)
        
        await callback.message.answer_document(
            document=input_file,
            caption=f"<b>{event.title}</b>\nРегистраций: {len(registrations)}",
            reply_markup=get_admin_export_back_keyboard(),
            parse_mode="HTML"
        )
        
    await callback.answer()


# ==========================================
# ОБРАТНАЯ СВЯЗЬ (ПРОСМОТР АДМИНИСТРАТОРОМ)
# ==========================================

class AdminFeedbackStates(StatesGroup):
    viewing = State()
    waiting_for_reply = State()


async def render_feedback_message(message: Message, feedback: FeedbackMessage, index: int, total: int):
    text = (
        f"💬 <b>Обращение {index + 1} из {total}</b>\n"
        f"<b>Отправитель:</b> <a href=\"tg://user?id={feedback.user_id}\">{feedback.full_name}</a> "
        f"({f'@{feedback.username}' if feedback.username else 'нет юзернейма'})\n"
        f"<b>ID:</b> <code>{feedback.user_id}</code>\n"
        f"<b>Дата отправки:</b> {feedback.created_at}\n\n"
        f"<b>Текст обращения:</b>\n{feedback.text}"
    )
    await message.edit_text(
        text,
        reply_markup=get_feedback_navigation_keyboard(feedback.user_id),
        parse_mode="HTML",
        disable_web_page_preview=True
    )


@router.callback_query(F.data == "admin_view_feedback")
async def process_admin_view_feedback(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        query = select(FeedbackMessage).order_by(FeedbackMessage.created_at.asc())
        result = await session.execute(query)
        feedbacks = result.scalars().all()

        if not feedbacks:
            await callback.message.answer(
                "Сообщений обратной связи пока нет.",
                reply_markup=get_cancel_keyboard(show_back=False)
            )
            await callback.answer()
            return

        total = len(feedbacks)
        latest_idx = total - 1
        await state.set_state(AdminFeedbackStates.viewing)
        await state.update_data(current_feedback_index=latest_idx)
        
        await render_feedback_message(callback.message, feedbacks[latest_idx], latest_idx, total)
        await callback.answer()


@router.callback_query(F.data == "feedback_nav_prev", AdminFeedbackStates.viewing)
async def process_feedback_nav_prev(callback: CallbackQuery, state: FSMContext):
    async with async_session() as session:
        query = select(FeedbackMessage).order_by(FeedbackMessage.created_at.asc())
        result = await session.execute(query)
        feedbacks = result.scalars().all()

        if not feedbacks:
            await state.clear()
            await callback.message.edit_text("Сообщений обратной связи пока нет.", reply_markup=get_cancel_keyboard(show_back=False))
            await callback.answer()
            return

        data = await state.get_data()
        current_idx = data.get("current_feedback_index", 0)
        total = len(feedbacks)
        
        new_idx = (current_idx - 1) % total
        await state.update_data(current_feedback_index=new_idx)
        await render_feedback_message(callback.message, feedbacks[new_idx], new_idx, total)
        await callback.answer()


@router.callback_query(F.data == "feedback_nav_next", AdminFeedbackStates.viewing)
async def process_feedback_nav_next(callback: CallbackQuery, state: FSMContext):
    async with async_session() as session:
        query = select(FeedbackMessage).order_by(FeedbackMessage.created_at.asc())
        result = await session.execute(query)
        feedbacks = result.scalars().all()

        if not feedbacks:
            await state.clear()
            await callback.message.edit_text("Сообщений обратной связи пока нет.", reply_markup=get_cancel_keyboard(show_back=False))
            await callback.answer()
            return

        data = await state.get_data()
        current_idx = data.get("current_feedback_index", 0)
        total = len(feedbacks)
        
        new_idx = (current_idx + 1) % total
        await state.update_data(current_feedback_index=new_idx)
        await render_feedback_message(callback.message, feedbacks[new_idx], new_idx, total)
        await callback.answer()


@router.callback_query(F.data.startswith("admin_reply_feedback_"))
async def process_admin_reply_feedback(callback: CallbackQuery, state: FSMContext):
    user_id = int(callback.data.split("_")[3])
    
    async with async_session() as session:
        # First try to find the name in the feedback messages to get the exact sender name
        query = select(FeedbackMessage).where(FeedbackMessage.user_id == user_id).order_by(FeedbackMessage.created_at.desc()).limit(1)
        res = await session.execute(query)
        fm = res.scalar_one_or_none()
        if fm:
            name = fm.full_name
            user_link = f'<a href="tg://user?id={user_id}">{name}</a>'
        else:
            from database.models import User
            user = await session.get(User, user_id)
            if user:
                name = user.full_name or "Пользователь"
                user_link = f'<a href="tg://user?id={user_id}">{name}</a>'
            else:
                user_link = f'<a href="tg://user?id={user_id}">Пользователь {user_id}</a>'
                
    await state.set_state(AdminFeedbackStates.waiting_for_reply)
    await state.update_data(reply_target_user_id=user_id, reply_target_link=user_link)
    
    await callback.message.answer(
        f"введите текст ответа на сообщение пользователя {user_link}",
        reply_markup=get_cancel_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()


@router.message(AdminFeedbackStates.waiting_for_reply)
async def process_admin_reply_message(message: Message, state: FSMContext):
    reply_text = message.text
    if not reply_text:
        await message.answer("Пожалуйста, введите текстовый ответ:")
        return
        
    data = await state.get_data()
    target_user_id = data.get("reply_target_user_id")
    target_link = data.get("reply_target_link")
    
    bot = message.bot
    user_notification = (
        "Новый ответ от поддержки бота <b>Эксклюзивно: Креативный хаб НИУ ВШЭ</b>:\n\n"
        f"{reply_text}"
    )
    
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    user_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Назад", callback_data="back_to_main")]
    ])
    
    try:
        await bot.send_message(
            chat_id=target_user_id,
            text=user_notification,
            parse_mode="HTML",
            reply_markup=user_kb
        )
        await message.answer(
            f"Ответ успешно отправлен пользователю {target_link}.",
            parse_mode="HTML",
            reply_markup=get_admin_main_keyboard()
        )
    except Exception as e:
        await message.answer(
            f"Не удалось отправить сообщение пользователю {target_link}. Ошибка: {e}",
            parse_mode="HTML",
            reply_markup=get_admin_main_keyboard()
        )
        
    await state.clear()


# ----- Добавление партнерского мероприятия -----

@router.callback_query(F.data == "admin_add_partner_event")
async def process_admin_add_partner_event(callback: CallbackQuery, state: FSMContext):
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return
        await state.set_state(PartnerEventForm.title)
        await callback.message.answer(
            "Вопрос 1: Название партнерского мероприятия",
            reply_markup=get_cancel_keyboard()
        )
    await callback.answer()


@router.message(PartnerEventForm.title)
async def process_add_partner_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text.strip())
    await state.set_state(PartnerEventForm.date)
    await message.answer(
        "Вопрос 2: Дата проведения (в формате ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_date")
    )


@router.callback_query(F.data == "back_partner_date", PartnerEventForm.date)
async def process_back_partner_date(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PartnerEventForm.title)
    await callback.message.answer(
        "Вопрос 1: Название партнерского мероприятия",
        reply_markup=get_cancel_keyboard()
    )
    await callback.answer()


@router.message(PartnerEventForm.date)
async def process_add_partner_date(message: Message, state: FSMContext):
    if not validate_date_format(message.text):
        await message.answer(
            "Неверный формат даты, попробуйте еще раз. Ожидается формат ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_date")
        )
        return
        
    await state.update_data(date=message.text.strip())
    await state.set_state(PartnerEventForm.description)
    await message.answer(
        "Вопрос 3: Текст описания",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_description")
    )


@router.callback_query(F.data == "back_partner_description", PartnerEventForm.description)
async def process_back_partner_description(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PartnerEventForm.date)
    await callback.message.answer(
        "Вопрос 2: Дата проведения (в формате ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_date")
    )
    await callback.answer()


@router.message(PartnerEventForm.description)
async def process_add_partner_description(message: Message, state: FSMContext):
    await state.update_data(description=message.html_text.strip())
    await state.set_state(PartnerEventForm.link)
    await message.answer(
        "Вопрос 4: Ссылка на подробную информацию",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_link")
    )


@router.callback_query(F.data == "back_partner_link", PartnerEventForm.link)
async def process_back_partner_link(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PartnerEventForm.description)
    await callback.message.answer(
        "Вопрос 3: Текст описания",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_description")
    )
    await callback.answer()


@router.message(PartnerEventForm.link)
async def process_add_partner_link(message: Message, state: FSMContext):
    link = message.text.strip()
    if not (link.startswith("http://") or link.startswith("https://")):
        await message.answer(
            "Некорректная ссылка, введите еще раз (должна начинаться с http:// или https://):",
            reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_partner_link")
        )
        return
        
    data = await state.get_data()
    async with async_session() as session:
        new_partner = PartnerEvent(
            title=data["title"],
            date=data["date"],
            description=data["description"],
            link=link
        )
        session.add(new_partner)
        await session.commit()
        
    await state.clear()
    await message.answer(
        "Партнерское мероприятие успешно добавлено!",
        reply_markup=get_admin_events_keyboard()
    )


# ----- Удаление партнерского мероприятия -----

@router.callback_query(F.data == "admin_del_partner_event_list")
async def process_admin_del_partner_event_list(callback: CallbackQuery):
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        query = select(PartnerEvent)
        result = await session.execute(query)
        events = result.scalars().all()
        
        today = datetime.now().date()
        active_partners = []
        for pe in events:
            try:
                if "-" in pe.date:
                    end_date_str = pe.date.split("-")[1].strip()
                else:
                    end_date_str = pe.date.strip()
                end_date = datetime.strptime(end_date_str, "%d.%m.%Y").date()
            except Exception:
                end_date = today
                
            if today <= end_date:
                active_partners.append(pe)

        if not active_partners:
            await callback.message.answer(
                "Нет активных партнерских мероприятий для удаления.",
                reply_markup=get_admin_events_keyboard()
            )
        else:
            await callback.message.answer(
                "Выберите партнерское мероприятие для удаления:",
                reply_markup=get_admin_partner_events_keyboard(active_partners)
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_del_pevent_"))
async def process_admin_del_pevent(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[3])
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        pevent = await session.get(PartnerEvent, event_id)
        if pevent:
            await session.delete(pevent)
            await session.commit()
            await callback.message.answer(
                f"Партнерское мероприятие «{pevent.title}» успешно удалено.",
                reply_markup=get_admin_events_keyboard()
            )
        else:
            await callback.answer("Мероприятие партнеров не найдено.")
    await callback.answer()


# ==========================================
# УПРАВЛЕНИЕ СЕРИЯМИ МЕРОПРИЯТИЙ (ДВЕРИ И Т.Д.)
# ==========================================

@router.callback_query(F.data == "admin_edit_series")
async def process_admin_edit_series(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

        res = await session.execute(select(EventSeries))
        series_list = res.scalars().all()

        await callback.message.answer(
            "Какую серию мероприятий редактировать?",
            reply_markup=get_admin_series_list_keyboard(series_list)
        )
    await callback.answer()


@router.callback_query(F.data == "admin_create_series")
async def process_admin_create_series(callback: CallbackQuery, state: FSMContext):
    async with async_session() as session:
        if not await is_user_admin(callback.from_user.username, session):
            await callback.answer("У вас нет прав доступа к этому разделу.", show_alert=True)
            return

    await state.set_state(CreateSeriesForm.title)
    await callback.message.answer(
        "Вопрос 1: Введите название серии (например, «ДВЕРИ: Открытое ревью работ»)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="admin_edit_series")
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_series_edit_info_"))
async def process_admin_series_edit_info(callback: CallbackQuery, state: FSMContext):
    series_id = int(callback.data.split("_")[4])
    await state.set_state(CreateSeriesForm.title)
    await state.update_data(editing_series_id=series_id)
    await callback.message.answer(
        "Вопрос 1: Введите новое название серии",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_series_{series_id}")
    )
    await callback.answer()


@router.callback_query(F.data == "back_series_title", CreateSeriesForm.description)
async def process_back_series_title(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    editing_series_id = data.get("editing_series_id")
    back_cb = f"admin_select_series_{editing_series_id}" if editing_series_id else "admin_edit_series"
    await state.set_state(CreateSeriesForm.title)
    await callback.message.answer(
        "Вопрос 1: Введите название серии",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback=back_cb)
    )
    await callback.answer()


@router.message(CreateSeriesForm.title)
async def process_create_series_title(message: Message, state: FSMContext):
    title = message.text.strip() if message.text else ""
    if not title:
        await message.answer("Пожалуйста, введите текстовое название серии:", reply_markup=get_cancel_keyboard(show_back=True, back_callback="admin_edit_series"))
        return
    await state.update_data(title=title)
    await state.set_state(CreateSeriesForm.description)
    await message.answer(
        "Вопрос 2: Введите описание серии (сохраняется форматирование, полученное от вас)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_series_title")
    )


@router.callback_query(F.data == "back_series_desc", CreateSeriesForm.image)
async def process_back_series_desc(callback: CallbackQuery, state: FSMContext):
    await state.set_state(CreateSeriesForm.description)
    await callback.message.answer(
        "Вопрос 2: Введите описание серии (сохраняется форматирование, полученное от вас)",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_series_title")
    )
    await callback.answer()


@router.message(CreateSeriesForm.description)
async def process_create_series_description(message: Message, state: FSMContext):
    desc = message.html_text if message.html_text else message.text
    if not desc:
        await message.answer("Пожалуйста, введите текстовое описание серии:", reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_series_title"))
        return
    await state.update_data(description=desc)
    await state.set_state(CreateSeriesForm.image)
    await message.answer(
        "Вопрос 3: Пришлите картинку-афишу серии файлом (без сжатия) или обычной фотографией",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_series_desc")
    )


@router.message(CreateSeriesForm.image, F.photo | F.document)
async def process_create_series_image(message: Message, state: FSMContext):
    image_id = None
    if message.photo:
        image_id = message.photo[-1].file_id
    elif message.document and message.document.mime_type and message.document.mime_type.startswith("image/"):
        image_id = message.document.file_id

    if not image_id:
        await message.answer("Пожалуйста, пришлите графический файл или фото:", reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_series_desc"))
        return

    data = await state.get_data()
    editing_series_id = data.get("editing_series_id")

    async with async_session() as session:
        if editing_series_id:
            series = await session.get(EventSeries, editing_series_id)
            if series:
                series.title = data["title"]
                series.description = data["description"]
                series.image_id = image_id
                session.add(series)
                await session.commit()
                target_series_id = series.id
        else:
            new_series = EventSeries(
                title=data["title"],
                description=data["description"],
                image_id=image_id
            )
            session.add(new_series)
            await session.commit()
            target_series_id = new_series.id

    await state.clear()
    await message.answer(
        "Описание серии успешно сохранено!",
        reply_markup=get_admin_series_actions_keyboard(target_series_id)
    )


@router.callback_query(F.data.startswith("admin_select_series_"))
async def process_admin_select_series(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    series_id = int(callback.data.split("_")[3])
    async with async_session() as session:
        series = await session.get(EventSeries, series_id)
        if not series:
            await callback.answer("Серия не найдена.")
            return

        await callback.message.answer(
            f"Управление серией «<b>{series.title}</b>»\nЧто вы хотите сделать?",
            reply_markup=get_admin_series_actions_keyboard(series.id)
        )
    await callback.answer()


# ----- Кнопка 2: Добавить событие в серию -----

@router.callback_query(F.data.startswith("admin_series_add_event_"))
async def process_admin_series_add_event(callback: CallbackQuery, state: FSMContext):
    series_id = int(callback.data.split("_")[4])
    await state.set_state(AddSeriesEventForm.topic)
    await state.update_data(series_id=series_id)
    await callback.message.answer(
        "1. Укажите тематику (название события в серии, например: «Иллюстрация»):",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_series_{series_id}")
    )
    await callback.answer()


@router.callback_query(F.data == "back_sevent_topic", AddSeriesEventForm.date)
async def process_back_sevent_topic(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    series_id = data["series_id"]
    await state.set_state(AddSeriesEventForm.topic)
    await callback.message.answer(
        "1. Укажите тематику (название события в серии, например: «Иллюстрация»):",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_series_{series_id}")
    )
    await callback.answer()


@router.message(AddSeriesEventForm.topic)
async def process_series_event_topic(message: Message, state: FSMContext):
    data = await state.get_data()
    series_id = data["series_id"]
    topic = message.text.strip() if message.text else ""
    if not topic:
        await message.answer("Пожалуйста, введите тематику события:", reply_markup=get_cancel_keyboard(show_back=True, back_callback=f"admin_select_series_{series_id}"))
        return
    await state.update_data(topic=topic)
    await state.set_state(AddSeriesEventForm.date)
    await message.answer(
        "2. Укажите дату события в формате ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ:",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_sevent_topic")
    )


@router.callback_query(F.data == "back_sevent_date", AddSeriesEventForm.time)
async def process_back_sevent_date(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AddSeriesEventForm.date)
    await callback.message.answer(
        "2. Укажите дату события в формате ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ:",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_sevent_topic")
    )
    await callback.answer()


@router.message(AddSeriesEventForm.date)
async def process_series_event_date(message: Message, state: FSMContext):
    date_str = message.text.strip() if message.text else ""
    if not date_str:
        await message.answer("Пожалуйста, укажите дату события:", reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_sevent_topic"))
        return
    await state.update_data(date=date_str)
    await state.set_state(AddSeriesEventForm.time)
    await message.answer(
        "3. Введите время события в формате ЧЧ:ММ или ЧЧ:ММ-ЧЧ:ММ:",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_sevent_date")
    )


@router.callback_query(F.data == "back_sevent_time", AddSeriesEventForm.extra_text)
async def process_back_sevent_time(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AddSeriesEventForm.time)
    await callback.message.answer(
        "3. Введите время события в формате ЧЧ:ММ или ЧЧ:ММ-ЧЧ:ММ:",
        reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_sevent_date")
    )
    await callback.answer()


@router.message(AddSeriesEventForm.time)
async def process_series_event_time(message: Message, state: FSMContext):
    time_str = message.text.strip() if message.text else ""
    if not time_str:
        await message.answer("Пожалуйста, введите время события:", reply_markup=get_cancel_keyboard(show_back=True, back_callback="back_sevent_date"))
        return
    await state.update_data(time=time_str)
    await state.set_state(AddSeriesEventForm.extra_text)
    await message.answer(
        "4. Введите дополнительный текст при желании (сохраняется пользовательское форматирование):",
        reply_markup=get_skip_keyboard("skip_series_extra_text", show_back=True, back_callback="back_sevent_time")
    )


@router.callback_query(F.data == "skip_series_extra_text", AddSeriesEventForm.extra_text)
async def process_skip_series_extra_text(callback: CallbackQuery, state: FSMContext):
    await save_series_event(callback.message, state, extra_text="")
    await callback.answer()


@router.message(AddSeriesEventForm.extra_text)
async def process_series_event_extra_text(message: Message, state: FSMContext):
    extra_text = message.html_text if message.html_text else (message.text or "")
    await save_series_event(message, state, extra_text=extra_text)


async def save_series_event(message_or_msg, state: FSMContext, extra_text: str):
    data = await state.get_data()
    series_id = data["series_id"]

    async with async_session() as session:
        new_event = SeriesEvent(
            series_id=series_id,
            topic=data["topic"],
            date=data["date"],
            time=data["time"],
            extra_text=extra_text,
            is_deleted=0
        )
        session.add(new_event)
        await session.commit()

    await state.clear()
    await message_or_msg.answer(
        "Событие серии успешно добавлено!",
        reply_markup=get_admin_series_actions_keyboard(series_id)
    )


# ----- Кнопка 3: Удалить событие в серии -----

@router.callback_query(F.data.startswith("admin_series_del_event_list_"))
async def process_admin_series_del_event_list(callback: CallbackQuery):
    series_id = int(callback.data.split("_")[5])
    async with async_session() as session:
        res = await session.execute(
            select(SeriesEvent).where(SeriesEvent.series_id == series_id, SeriesEvent.is_deleted == 0)
        )
        events = res.scalars().all()

        if not events:
            await callback.message.answer(
                "В этой серии пока нет активных событий для удаления.",
                reply_markup=get_admin_series_actions_keyboard(series_id)
            )
        else:
            await callback.message.answer(
                "Выберите событие для удаления из серии:",
                reply_markup=get_series_events_select_keyboard(events, "admin_confirm_del_sevent", series_id)
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_confirm_del_sevent_"))
async def process_admin_confirm_del_sevent(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[4])
    async with async_session() as session:
        event = await session.get(SeriesEvent, event_id)
        if event:
            event.is_deleted = 1
            session.add(event)
            await session.commit()
            await callback.message.answer(
                f"Событие «{event.topic}» удалено из серии. Все сохраненные ранее заявки и базы сохраняются!",
                reply_markup=get_admin_series_actions_keyboard(event.series_id)
            )
        else:
            await callback.answer("Событие серии не найдено.")
    await callback.answer()


# ----- Кнопка 5: Редактировать анкету регистрации (Конструктор) -----

@router.callback_query(F.data.startswith("admin_series_edit_form_"))
async def process_admin_series_edit_form(callback: CallbackQuery, state: FSMContext):
    series_id = int(callback.data.split("_")[4])
    await state.set_state(SeriesQuestionnaireForm.asking_question)
    await state.update_data(series_id=series_id, questions=[])

    await callback.message.answer(
        "<b>Конструктор анкеты серии</b>\n\n"
        "Вопрос 1. Введите текст вопроса для анкеты с итоговым форматированием:",
        reply_markup=get_questionnaire_loop_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data == "admin_q_back", SeriesQuestionnaireForm.asking_question)
async def process_admin_q_back(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    questions = data.get("questions", [])
    if questions:
        questions.pop()
        await state.update_data(questions=questions)
    
    q_num = len(questions) + 1
    await callback.message.answer(
        f"Вопрос {q_num}. Введите текст вопроса для анкеты с итоговым форматированием:",
        reply_markup=get_questionnaire_loop_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()


@router.message(SeriesQuestionnaireForm.asking_question)
async def process_admin_q_message(message: Message, state: FSMContext):
    q_text = message.html_text if message.html_text else message.text
    if not q_text:
        await message.answer("Пожалуйста, введите текстовый вопрос:", reply_markup=get_questionnaire_loop_keyboard())
        return

    data = await state.get_data()
    questions = data.get("questions", [])
    questions.append(q_text)
    await state.update_data(questions=questions)

    q_num = len(questions) + 1
    await message.answer(
        f"Вопрос {q_num}. Введите текст вопроса для анкеты с итоговым форматированием:",
        reply_markup=get_questionnaire_loop_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(F.data == "admin_q_finish", SeriesQuestionnaireForm.asking_question)
async def process_admin_q_finish(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    series_id = data["series_id"]
    questions = data.get("questions", [])

    async with async_session() as session:
        # Удаляем старые вопросы серии и записываем новый список вопросов
        await session.execute(
            SeriesQuestion.__table__.delete().where(SeriesQuestion.series_id == series_id)
        )
        for idx, q_text in enumerate(questions, start=1):
            sq = SeriesQuestion(
                series_id=series_id,
                question_order=idx,
                question_text=q_text
            )
            session.add(sq)
        await session.commit()

    await state.clear()
    
    summary = "\n".join([f"{i}. {q}" for i, q in enumerate(questions, 1)]) if questions else "Вопросы отсутствуют."
    await callback.message.answer(
        f"Анкета серии успешно создана!\n\n<b>Список вопросов:</b>\n{summary}",
        reply_markup=get_admin_series_actions_keyboard(series_id),
        parse_mode="HTML"
    )
    await callback.answer()


# ----- Кнопка 1: Выгрузить список регистраций серии -----

@router.callback_query(F.data.startswith("admin_series_export_"))
async def process_admin_series_export(callback: CallbackQuery):
    series_id = int(callback.data.split("_")[3])
    async with async_session() as session:
        res = await session.execute(
            select(SeriesEvent).where(SeriesEvent.series_id == series_id)
        )
        events = res.scalars().all()

        if not events:
            await callback.message.answer(
                "В этой серии пока нет событий для выгрузки заявок.",
                reply_markup=get_admin_export_type_keyboard()
            )
        else:
            await callback.message.answer(
                "Выберите событие серии для выгрузки Excel-таблицы заявок:",
                reply_markup=get_series_events_select_keyboard(events, "admin_export_sevent", series_id, back_callback="admin_export_registrations")
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_export_sevent_"))
async def process_admin_export_sevent(callback: CallbackQuery):
    event_id = int(callback.data.split("_")[3])
    import openpyxl
    import io

    async with async_session() as session:
        sevent = await session.get(SeriesEvent, event_id)
        if not sevent:
            await callback.answer("Событие серии не найдено.")
            return

        res = await session.execute(
            select(SeriesApplication).where(SeriesApplication.series_event_id == event_id).order_by(SeriesApplication.created_at.asc())
        )
        apps = res.scalars().all()

        if not apps:
            await callback.message.answer(
                f"На событие «{sevent.topic}» пока нет поданных заявок.",
                reply_markup=get_admin_export_back_keyboard()
            )
            await callback.answer()
            return

        # Динамически собираем все уникальные вопросы, встречающиеся в ответах пользователей
        header_questions = []
        for app in apps:
            if isinstance(app.answers, dict):
                for q_title in app.answers.keys():
                    if q_title not in header_questions:
                        header_questions.append(q_title)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заявки серии"

        # Заголовки столбцов
        headers = ["Telegram ID", "Никнейм", "Дата и время подачи"] + header_questions
        ws.append(headers)

        for app in apps:
            username_str = f"@{app.username}" if app.username else "нет юзернейма"
            answers_dict = app.answers if isinstance(app.answers, dict) else {}
            row = [app.user_id, username_str, app.created_at]
            for q_title in header_questions:
                row.append(answers_dict.get(q_title, ""))
            ws.append(row)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"applications_series_event_{event_id}.xlsx"
        input_file = BufferedInputFile(stream.read(), filename=filename)

        await callback.message.answer_document(
            document=input_file,
            caption=f"Выгрузка заявок по событию «{sevent.topic}» ({len(apps)} шт.)",
            reply_markup=get_admin_export_back_keyboard()
        )
    await callback.answer()


# ----- ОТОБРАТЬ ПОБЕДИТЕЛЕЙ -----

def get_fio_from_answers(answers: dict, username: str = None) -> str:
    if isinstance(answers, dict):
        for q, a in answers.items():
            if any(k in q.lower() for k in ["фио", "имя", "фамилия"]):
                return str(a).strip()
        first_val = next(iter(answers.values()), None)
        if first_val and len(str(first_val)) < 50:
            return str(first_val).strip()
    return f"@{username}" if username else "Участник"


def get_work_link_from_answers(answers: dict) -> str | None:
    if isinstance(answers, dict):
        for q, a in answers.items():
            a_str = str(a).strip()
            if any(k in q.lower() for k in ["портфолио", "работы", "работа", "ссылка", "диск"]):
                return a_str
            if a_str.startswith("http://") or a_str.startswith("https://"):
                return a_str
    return None


@router.callback_query(F.data.startswith("admin_series_winners_"))
async def process_admin_series_winners(callback: CallbackQuery, state: FSMContext):
    series_id = int(callback.data.split("_")[3])
    async with async_session() as session:
        res = await session.execute(
            select(SeriesEvent).where(SeriesEvent.series_id == series_id, SeriesEvent.is_deleted == 0)
        )
        events = res.scalars().all()

        if not events:
            await callback.message.answer(
                "В этой серии пока нет активных событий для отбора победителей.",
                reply_markup=get_admin_series_actions_keyboard(series_id)
            )
        else:
            await callback.message.answer(
                "Выберите событие серии для отбора участников:",
                reply_markup=get_series_events_select_keyboard(events, "admin_winner_sevent", series_id)
            )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_winner_sevent_"))
async def process_admin_winner_sevent(callback: CallbackQuery, state: FSMContext):
    event_id = int(callback.data.split("_")[3])
    async with async_session() as session:
        sevent = await session.get(SeriesEvent, event_id)
        if not sevent:
            await callback.answer("Событие серии не найдено.")
            return

        res = await session.execute(
            select(SeriesApplication).where(SeriesApplication.series_event_id == event_id).order_by(SeriesApplication.created_at.asc())
        )
        apps = res.scalars().all()

        if not apps:
            await callback.message.answer(
                f"На событие «{sevent.topic}» пока нет поданных заявок.",
                reply_markup=get_admin_series_actions_keyboard(sevent.series_id)
            )
            await callback.answer()
            return

        selected_ids = set()
        await state.set_state(SelectWinnersForm.selected_app_ids)
        await state.update_data(
            series_id=sevent.series_id,
            series_event_id=event_id,
            selected_app_ids=list(selected_ids)
        )

        apps_with_names = []
        for app in apps:
            fio = get_fio_from_answers(app.answers, app.username)
            username_part = f"@{app.username}" if app.username else f"id:{app.user_id}"
            display_name = f"{username_part} ({fio})"
            apps_with_names.append((app.id, display_name))

        await callback.message.answer(
            "Прокликайте пользователей, которые получат сообщение о прохождении отбора:",
            reply_markup=get_winners_checkbox_keyboard(apps_with_names, selected_ids, sevent.series_id)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_toggle_winner_"), SelectWinnersForm.selected_app_ids)
async def process_admin_toggle_winner(callback: CallbackQuery, state: FSMContext):
    app_id = int(callback.data.split("_")[3])
    data = await state.get_data()
    selected_ids = set(data.get("selected_app_ids", []))

    if app_id in selected_ids:
        selected_ids.remove(app_id)
    else:
        selected_ids.add(app_id)

    await state.update_data(selected_app_ids=list(selected_ids))

    series_event_id = data["series_event_id"]
    series_id = data["series_id"]

    async with async_session() as session:
        res = await session.execute(
            select(SeriesApplication).where(SeriesApplication.series_event_id == series_event_id).order_by(SeriesApplication.created_at.asc())
        )
        apps = res.scalars().all()

        apps_with_names = []
        for app in apps:
            fio = get_fio_from_answers(app.answers, app.username)
            username_part = f"@{app.username}" if app.username else f"id:{app.user_id}"
            display_name = f"{username_part} ({fio})"
            apps_with_names.append((app.id, display_name))

        try:
            await callback.message.edit_reply_markup(
                reply_markup=get_winners_checkbox_keyboard(apps_with_names, selected_ids, series_id)
            )
        except Exception:
            pass
    await callback.answer()


@router.callback_query(F.data == "admin_winner_back_to_checkboxes")
async def process_admin_winner_back_to_checkboxes(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected_ids = set(data.get("selected_app_ids", []))
    series_event_id = data["series_event_id"]
    series_id = data["series_id"]

    await state.set_state(SelectWinnersForm.selected_app_ids)

    async with async_session() as session:
        res = await session.execute(
            select(SeriesApplication).where(SeriesApplication.series_event_id == series_event_id).order_by(SeriesApplication.created_at.asc())
        )
        apps = res.scalars().all()

        apps_with_names = []
        for app in apps:
            fio = get_fio_from_answers(app.answers, app.username)
            username_part = f"@{app.username}" if app.username else f"id:{app.user_id}"
            display_name = f"{username_part} ({fio})"
            apps_with_names.append((app.id, display_name))

        await callback.message.answer(
            "Прокликайте пользователей, которые получат сообщение о прохождении отбора:",
            reply_markup=get_winners_checkbox_keyboard(apps_with_names, selected_ids, series_id)
        )
    await callback.answer()


@router.callback_query(F.data == "admin_winner_save_selection", SelectWinnersForm.selected_app_ids)
async def process_admin_winner_save_selection(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected_ids = data.get("selected_app_ids", [])

    if not selected_ids:
        await callback.answer("Выберите хотя бы одного пользователя для прохождения отбора.", show_alert=True)
        return

    await state.set_state(SelectWinnersForm.confirming)

    async with async_session() as session:
        lines = ["Проверьте список пользователей, которые получат сообщение о прохождении отбора:\n"]
        for idx, app_id in enumerate(selected_ids, 1):
            app = await session.get(SeriesApplication, app_id)
            if app:
                fio = get_fio_from_answers(app.answers, app.username)
                username_str = f"@{app.username}" if app.username else f"id:{app.user_id}"
                work_link = get_work_link_from_answers(app.answers)
                
                lines.append(f"{idx}. {username_str} ({fio})")
                if work_link:
                    if work_link.startswith("http://") or work_link.startswith("https://"):
                        lines.append(f'   <a href="{work_link}">Ссылка на работы</a>')
                    else:
                        lines.append(f"   Ссылка на работы: {work_link}")

        summary_text = "\n".join(lines)
        await callback.message.answer(
            summary_text,
            reply_markup=get_winners_confirm_keyboard(),
            parse_mode="HTML",
            disable_web_page_preview=True
        )
    await callback.answer()


@router.callback_query(F.data == "admin_winner_back_to_confirm", SelectWinnersForm.entering_extra_text)
async def process_admin_winner_back_to_confirm(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected_ids = data.get("selected_app_ids", [])
    await state.set_state(SelectWinnersForm.confirming)

    async with async_session() as session:
        lines = ["Проверьте список пользователей, которые получат сообщение о прохождении отбора:\n"]
        for idx, app_id in enumerate(selected_ids, 1):
            app = await session.get(SeriesApplication, app_id)
            if app:
                fio = get_fio_from_answers(app.answers, app.username)
                username_str = f"@{app.username}" if app.username else f"id:{app.user_id}"
                work_link = get_work_link_from_answers(app.answers)
                
                lines.append(f"{idx}. {username_str} ({fio})")
                if work_link:
                    if work_link.startswith("http://") or work_link.startswith("https://"):
                        lines.append(f'   <a href="{work_link}">Ссылка на работы</a>')
                    else:
                        lines.append(f"   Ссылка на работы: {work_link}")

        summary_text = "\n".join(lines)
        await callback.message.answer(
            summary_text,
            reply_markup=get_winners_confirm_keyboard(),
            parse_mode="HTML",
            disable_web_page_preview=True
        )
    await callback.answer()


@router.callback_query(F.data == "admin_winner_proceed", SelectWinnersForm.confirming)
async def process_admin_winner_proceed(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    series_event_id = data["series_event_id"]
    selected_ids = data.get("selected_app_ids", [])

    await state.set_state(SelectWinnersForm.entering_extra_text)

    async with async_session() as session:
        sevent = await session.get(SeriesEvent, series_event_id)
        series = await session.get(EventSeries, sevent.series_id) if sevent else None

        fio_sample = "{ФИО из анкеты}"
        if selected_ids:
            first_app = await session.get(SeriesApplication, selected_ids[0])
            if first_app:
                fio_sample = get_fio_from_answers(first_app.answers, first_app.username)

        topic_str = sevent.topic if sevent else "Мероприятие"
        series_title = series.title if series else "Серия"

        preview_text = (
            "Пользователь получит сообщение:\n\n"
            f"<b>{fio_sample}, ваша заявка одобрена!</b>\n\n"
            f"Поздравляем, жюри отобрало вашу заявку на участие в мероприятии: <b>{topic_str}</b> серии <b>{series_title}</b>.\n\n"
            "Хотите что-то добавить? Введите текст ниже."
        )

        await callback.message.answer(
            preview_text,
            reply_markup=get_cancel_keyboard(show_back=True, back_callback="admin_winner_back_to_confirm"),
            parse_mode="HTML"
        )
    await callback.answer()


@router.message(SelectWinnersForm.entering_extra_text)
async def process_admin_winner_extra_text(message: Message, state: FSMContext):
    extra_text = message.html_text if message.html_text else (message.text or "")
    await state.update_data(extra_text=extra_text)

    data = await state.get_data()
    series_event_id = data["series_event_id"]
    selected_ids = data.get("selected_app_ids", [])

    async with async_session() as session:
        sevent = await session.get(SeriesEvent, series_event_id)
        series = await session.get(EventSeries, sevent.series_id) if sevent else None

        fio_sample = "{ФИО из анкеты}"
        if selected_ids:
            first_app = await session.get(SeriesApplication, selected_ids[0])
            if first_app:
                fio_sample = get_fio_from_answers(first_app.answers, first_app.username)

        topic_str = sevent.topic if sevent else "Мероприятие"
        series_title = series.title if series else "Серия"

        preview_text = (
            "Пользователь получит сообщение:\n\n"
            f"<b>{fio_sample}, ваша заявка одобрена!</b>\n\n"
            f"Поздравляем, жюри отобрало вашу заявку на участие в мероприятии: <b>{topic_str}</b> серии <b>{series_title}</b>.\n\n"
            f"{extra_text}"
        )

        await message.answer(
            preview_text,
            reply_markup=get_winners_dispatch_keyboard(),
            parse_mode="HTML"
        )


@router.callback_query(F.data == "admin_winner_dispatch")
async def process_admin_winner_dispatch(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected_ids = data.get("selected_app_ids", [])
    series_event_id = data.get("series_event_id")
    extra_text = data.get("extra_text", "")

    bot = getattr(callback.message, 'bot', None) or getattr(callback, 'bot', None)

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    user_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="← К списку", callback_data="back_to_main")]
    ])

    sent_count = 0
    async with async_session() as session:
        sevent = await session.get(SeriesEvent, series_event_id)
        series = await session.get(EventSeries, sevent.series_id) if sevent else None
        topic_str = sevent.topic if sevent else "Мероприятие"
        series_title = series.title if series else "Серия"

        for app_id in selected_ids:
            app = await session.get(SeriesApplication, app_id)
            if app:
                fio = get_fio_from_answers(app.answers, app.username)
                user_msg_text = (
                    f"<b>{fio}, ваша заявка одобрена!</b>\n\n"
                    f"Поздравляем, жюри отобрало вашу заявку на участие в мероприятии: <b>{topic_str}</b> серии <b>{series_title}</b>."
                )
                if extra_text:
                    user_msg_text += f"\n\n{extra_text}"

                if bot:
                    try:
                        await bot.send_message(
                            chat_id=app.user_id,
                            text=user_msg_text,
                            parse_mode="HTML",
                            reply_markup=user_kb
                        )
                        sent_count += 1
                    except Exception:
                        pass
                else:
                    sent_count += 1

    await state.clear()
    await callback.message.answer(
        f"Сообщение о прохождении отбора успешно отправлено выбранным участникам ({sent_count} чел.)!",
        reply_markup=get_admin_series_actions_keyboard(sevent.series_id if sevent else 1)
    )
    await callback.answer()




