import asyncio
import os
import sys
import json
import traceback
from datetime import datetime, timedelta
import config

# Use test database
config.DATABASE_URL = "sqlite+aiosqlite:///test_creative_hub.db"

# Reconfigure stdout to print utf-8
sys.stdout.reconfigure(encoding='utf-8')

from sqlalchemy import select, delete, func
from database.db import init_db, async_session
from database.models import User, Event, Registration, Raffle, Admin, SystemTag, FeedbackMessage, ReadPostMaterials, PartnerEvent

# --- Mock Classes for Aiogram ---
class DummyUser:
    def __init__(self, id, username):
        self.id = id
        self.username = username
        self.first_name = "Dummy"
        self.last_name = "User"

    @property
    def full_name(self):
        return f"{self.first_name} {self.last_name}"

class DummyMessage:
    def __init__(self, from_user_id, from_username, text=""):
        self.from_user = DummyUser(from_user_id, from_username)
        self.text = text
        self.html_text = text
        self.sent_messages = []
        self.deleted = False
        self.document = None

    async def answer(self, text, reply_markup=None, parse_mode=None, disable_web_page_preview=False):
        self.sent_messages.append((text, reply_markup))
        return self

    async def answer_photo(self, photo, caption, reply_markup=None, parse_mode=None):
        self.sent_messages.append((caption, reply_markup))
        return self

    async def delete(self):
        self.deleted = True

    async def edit_text(self, text, reply_markup=None, *args, **kwargs):
        self.sent_messages.append((text, reply_markup))
        return self

    async def edit_reply_markup(self, reply_markup=None):
        self.sent_messages.append(("", reply_markup))

    async def answer_document(self, document, caption=None, reply_markup=None, parse_mode=None):
        self.sent_messages.append((caption or "", document, reply_markup))
        return self

class DummyCallbackQuery:
    def __init__(self, data, from_user_id, from_username, message):
        self.data = data
        self.from_user = DummyUser(from_user_id, from_username)
        self.message = message
        self.answered = False
        self.answer_text = ""

    async def answer(self, text="", show_alert=False):
        self.answered = True
        self.answer_text = text

class DummyState:
    def __init__(self):
        self.state = None
        self.data = {}

    async def set_state(self, state):
        self.state = state

    async def get_data(self):
        return self.data

    async def update_data(self, **kwargs):
        self.data.update(kwargs)

    async def clear(self):
        self.state = None
        self.data = {}

class DummyBot:
    def __init__(self):
        self.sent_messages = []

    async def send_message(self, chat_id, text, parse_mode=None, disable_web_page_preview=False, reply_markup=None):
        self.sent_messages.append((chat_id, text, reply_markup))
        return None

    async def send_photo(self, chat_id, photo, caption=None, reply_markup=None, parse_mode=None):
        self.sent_messages.append((chat_id, caption, reply_markup))
        return None

# --- Import Bot Modules ---
import handlers
import admin_handlers
import scheduler
from keyboards import (
    get_main_menu_keyboard, get_events_list_keyboard, get_event_detail_keyboard,
    get_after_registration_keyboard, get_preferences_menu_keyboard, get_checkbox_keyboard,
    get_after_save_tags_keyboard, get_archive_tags_keyboard, get_archive_events_keyboard,
    get_archive_detail_keyboard, get_raffle_detail_keyboard, get_partners_events_keyboard
)
from admin_keyboards import (
    get_admin_main_keyboard, get_admin_tags_keyboard,
    get_admin_rights_keyboard, get_admin_raffles_keyboard, get_raffle_sections_keyboard,
    get_admin_events_keyboard, get_admin_delete_select_keyboard, get_post_mats_tags_keyboard,
    get_admin_export_events_keyboard, get_admin_export_archive_events_keyboard
)

async def test_suite():
    # Force fresh database recreation
    if os.path.exists("test_creative_hub.db"):
        try:
            os.remove("test_creative_hub.db")
        except Exception:
            pass
            
    print("Initializing test database...")
    await init_db()

    async with async_session() as session:
        # Clear tables
        await session.execute(User.__table__.delete())
        await session.execute(Event.__table__.delete())
        await session.execute(Registration.__table__.delete())
        await session.execute(Raffle.__table__.delete())
        await session.execute(Admin.__table__.delete())
        await session.execute(SystemTag.__table__.delete())
        await session.commit()

        # Seed System Tags
        for tag in config.DEFAULT_TAGS:
            session.add(SystemTag(name=tag))
        await session.commit()

        # ----------------------------------------------------
        # 1. Test Keyboards
        # ----------------------------------------------------
        print("Testing keyboard generation functions...")
        k1 = get_main_menu_keyboard(is_admin=True, raffle_count=5)
        k2 = get_checkbox_keyboard({"Digital_Дизайн": True}, "tags")
        k3 = get_admin_main_keyboard()
        k4 = get_post_mats_tags_keyboard(config.DEFAULT_TAGS, is_delete=False)
        assert len(k1.inline_keyboard) > 0
        assert len(k2.inline_keyboard) > 0
        assert len(k3.inline_keyboard) > 0
        assert len(k4.inline_keyboard) > 0

        # ----------------------------------------------------
        # 2. Command /start and User Onboarding & FSM Registration
        # ----------------------------------------------------
        print("Testing /start handler and user onboarding...")
        msg = DummyMessage(from_user_id=123, from_username="john_doe")
        state = DummyState()
        await handlers.cmd_start(msg, state)
        
        # Verify user created in DB
        user = await session.get(User, 123)
        assert user is not None
        assert user.username == "john_doe"
        assert user.tags_preferences["Digital_Дизайн"] is True
        assert user.is_registered is False

        # Run FSM registration flow
        assert state.state == handlers.RegistrationStates.accept_agreement
        cb_accept = DummyCallbackQuery("reg_accept", 123, "john_doe", msg)
        await handlers.process_reg_accept(cb_accept, state)
        
        assert state.state == handlers.RegistrationStates.full_name
        await handlers.process_reg_full_name(DummyMessage(123, "john_doe", "Кузнецов Пётр"), state)
        assert state.state == handlers.RegistrationStates.email

        # Email validation failure
        await handlers.process_reg_email(DummyMessage(123, "john_doe", "invalid-email"), state)
        assert state.state == handlers.RegistrationStates.email

        # Valid email
        await handlers.process_reg_email(DummyMessage(123, "john_doe", "ivanov@example.com"), state)
        assert state.state == handlers.RegistrationStates.phone

        # Skip phone step to complete registration
        cb_skip_phone = DummyCallbackQuery("reg_skip_phone", 123, "john_doe", msg)
        await handlers.process_reg_skip_phone(cb_skip_phone, state)

        # Verify registration values
        await session.refresh(user)
        assert user.is_registered is True
        assert user.full_name == "Кузнецов Пётр"
        assert user.email == "ivanov@example.com"
        assert user.phone == "-"


        # ----------------------------------------------------
        # 3. Preference Toggling and Saving
        # ----------------------------------------------------
        print("Testing preference checkbox toggling...")
        cb_pref = DummyCallbackQuery("pref_tags", 123, "john_doe", msg)
        await handlers.process_pref_tags(cb_pref, state)

        # Toggle Digital_Дизайн (index 1)
        cb_toggle = DummyCallbackQuery("cb_tags_1", 123, "john_doe", msg)
        await handlers.process_checkbox_toggle(cb_toggle, state)

        # Toggle another tag to test multiple interactions
        cb_toggle2 = DummyCallbackQuery("cb_tags_3", 123, "john_doe", msg)
        await handlers.process_checkbox_toggle(cb_toggle2, state)

        # Save preferences
        cb_save = DummyCallbackQuery("save_tags", 123, "john_doe", msg)
        await handlers.process_save_preferences(cb_save, state)

        # Reload and check
        await session.refresh(user)
        assert user.tags_preferences["Digital_Дизайн"] is False
        assert user.tags_preferences["КреативныйБизнес"] is False

        # ----------------------------------------------------
        # 4. Events Lifecycle & Preferences Filtering
        # ----------------------------------------------------
        print("Testing events details & preferences filtering...")
        # Create an event matching user's disabled tag
        event_dis = Event(
            id=1,
            title="Digital design stuff",
            date="10.06.2026",
            time="12:00",
            address="HSE",
            tags=["Digital_Дизайн"]
        )
        # Create an event matching user's active tag
        event_act = Event(
            id=2,
            title="Science talk",
            date="11.06.2026",
            time="14:00",
            address="HSE Science Hall",
            tags=["Наука"]
        )
        session.add_all([event_dis, event_act])
        await session.commit()

        # Query events list
        cb_events = DummyCallbackQuery("btn_events_info", 123, "john_doe", msg)
        await handlers.process_events_info(cb_events, state)
        
        # Check that only event 2 is shown
        kb = msg.sent_messages[-1][1]
        btn_texts = [btn.text for row in kb.inline_keyboard for btn in row]
        assert "Science talk" in btn_texts
        assert "Digital design stuff" not in btn_texts

        # ----------------------------------------------------
        # 5. Registrations (Очно / Удаленно / ... ) & Interception
        # ----------------------------------------------------
        print("Testing event registration status handling...")
        cb_reg = DummyCallbackQuery("reg_status_2_очно", 123, "john_doe", msg)
        await handlers.process_registration_status(cb_reg, state)

        reg = await session.get(Registration, (123, 2))
        assert reg is not None
        assert reg.status == "очно"
        assert reg.registration_date is not None

        # Test registration interception for unregistered user
        unreg_user = User(telegram_id=456, username="unreg_test")
        session.add(unreg_user)
        await session.commit()

        state2 = DummyState()
        cb_reg_unreg = DummyCallbackQuery("reg_status_2_удаленно", 456, "unreg_test", msg)
        await handlers.process_registration_status(cb_reg_unreg, state2)

        # Registration should be intercepted and FSM started
        assert state2.state == handlers.RegistrationStates.accept_agreement
        cb_accept2 = DummyCallbackQuery("reg_accept", 456, "unreg_test", msg)
        await handlers.process_reg_accept(cb_accept2, state2)

        assert state2.state == handlers.RegistrationStates.full_name
        reg_unreg_check = await session.get(Registration, (456, 2))
        assert reg_unreg_check is None

        # Complete registration for this user
        await handlers.process_reg_full_name(DummyMessage(456, "unreg_test", "Тестов Тест"), state2)
        await handlers.process_reg_email(DummyMessage(456, "unreg_test", "test@example.com"), state2)
        await handlers.process_reg_phone(DummyMessage(456, "unreg_test", "+79997777777"), state2)

        # Now registration should be automatically created after FSM finishes
        reg_unreg_check2 = await session.get(Registration, (456, 2))
        assert reg_unreg_check2 is not None
        assert reg_unreg_check2.status == "удаленно"
        assert reg_unreg_check2.registration_date is not None


        # Test change of status to "думаю" updates the registration status to "думаю" (remains in DB)
        cb_change_thinking = DummyCallbackQuery("reg_status_2_думаю", 456, "unreg_test", msg)
        await handlers.process_registration_status(cb_change_thinking, state2)
        session.expire_all()
        reg_unreg_thinking = await session.get(Registration, (456, 2))
        assert reg_unreg_thinking is not None
        assert reg_unreg_thinking.status == "думаю"

        # Test change of status to "не пойду" deletes the registration
        cb_change_nopgo = DummyCallbackQuery("reg_status_2_не пойду", 456, "unreg_test", msg)
        await handlers.process_registration_status(cb_change_nopgo, state2)
        session.expire_all()
        reg_unreg_deleted = await session.get(Registration, (456, 2))
        assert reg_unreg_deleted is None


        # ----------------------------------------------------
        # 6. Admin Authentication & Controls
        # ----------------------------------------------------
        print("Testing admin access controls...")
        normal_cb = DummyCallbackQuery("btn_admin", 123, "john_doe", msg)
        await admin_handlers.process_admin_menu(normal_cb, state)
        # Verify access denied alert was triggered
        assert normal_cb.answered is True
        assert "нет прав доступа" in normal_cb.answer_text.lower()

        # Admin user authentication (super admin configured in config)
        admin_cb = DummyCallbackQuery("btn_admin", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_menu(admin_cb, state)
        # Verify keyboard returned in admin_cb.message.sent_messages
        assert len(msg.sent_messages[-1][0]) > 0

        # ----------------------------------------------------
        # 7. Admin Tags CRUD
        # ----------------------------------------------------
        print("Testing admin tags CRUD...")
        # Create tag
        await admin_handlers.process_save_tag(DummyMessage(999, "ASaavedraA", "НовыйТестТаг"), state)
        # Confirm del tag with cascade validation
        del_cb = DummyCallbackQuery("admin_confirm_del_tag_НовыйТестТаг", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_confirm_del_tag(del_cb)

        # ----------------------------------------------------
        # 8. Admin Post-Materials CRUD
        # ----------------------------------------------------
        print("Testing admin post-materials addition/deletion...")
        
        # Test Scenario A: photographer is specified
        state_pm = DummyState()
        await state_pm.update_data(event_id=2)
        await state_pm.set_state(admin_handlers.PostMatsForm.photos_url)
        
        # Question 1: photos url
        pm_msg = DummyMessage(999, "ASaavedraA", "https://photos.url/123")
        await admin_handlers.process_add_pm_photos(pm_msg, state_pm)
        assert state_pm.state == admin_handlers.PostMatsForm.photographer_name
        
        # Question 2: photographer name
        photog_name_msg = DummyMessage(999, "ASaavedraA", "Иван Иванов")
        await admin_handlers.process_add_pm_photographer_name(photog_name_msg, state_pm)
        assert state_pm.state == admin_handlers.PostMatsForm.photographer_url
        
        # Question 3: photographer url
        photog_url_msg = DummyMessage(999, "ASaavedraA", "https://photographer.com")
        await admin_handlers.process_add_pm_photographer_url(photog_url_msg, state_pm)
        assert state_pm.state == admin_handlers.PostMatsForm.stream_record_url
        
        # Question 4: skip stream record
        skip_stream_cb = DummyCallbackQuery("skip_pm_stream", 999, "ASaavedraA", msg)
        await admin_handlers.process_skip_pm_stream(skip_stream_cb, state_pm)
        assert state_pm.state == admin_handlers.PostMatsForm.article_url
        
        # Question 5: skip article
        skip_art_cb = DummyCallbackQuery("skip_pm_article", 999, "ASaavedraA", msg)
        await admin_handlers.process_skip_pm_article(skip_art_cb, state_pm)
        assert state_pm.state == admin_handlers.PostMatsForm.presentations_url
        
        # Question 6: skip presentations
        skip_pres_cb = DummyCallbackQuery("skip_pm_presentations", 999, "ASaavedraA", msg)
        await admin_handlers.process_skip_pm_presentations(skip_pres_cb, state_pm)
        assert state_pm.state == admin_handlers.PostMatsForm.other_materials_url
        
        # Question 7: finish adding other materials
        finish_pm_cb = DummyCallbackQuery("finish_pm_adding", 999, "ASaavedraA", msg)
        await admin_handlers.process_finish_pm_adding_btn(finish_pm_cb, state_pm)
        assert state_pm.state is None # FSM finished
        
        # Verify db values for Event 2
        async with async_session() as test_session:
            ev2 = await test_session.get(Event, 2)
            assert ev2.photos_url == "https://photos.url/123"
            assert ev2.photographer_name == "Иван Иванов"
            assert ev2.photographer_url == "https://photographer.com"
            
        # Verify display formatting on Event 2 details screen
        detail_msg = DummyMessage(123, "john_doe")
        detail_cb = DummyCallbackQuery("arch_event_2", 123, "john_doe", detail_msg)
        await handlers.process_archive_event_detail(detail_cb)
        
        detail_text = detail_msg.sent_messages[-1][0]
        assert "💚 Фотограф: <a href=\"https://photographer.com\">Иван Иванов</a>" in detail_text
        
        # Test Scenario B: photographer is skipped
        state_pm_skip = DummyState()
        await state_pm_skip.update_data(event_id=2)
        await state_pm_skip.set_state(admin_handlers.PostMatsForm.photos_url)
        
        # Question 1: photos url
        await admin_handlers.process_add_pm_photos(pm_msg, state_pm_skip)
        
        # Question 2: skip photographer name
        skip_photog_cb = DummyCallbackQuery("skip_pm_photographer_name", 999, "ASaavedraA", msg)
        await admin_handlers.process_skip_pm_photographer_name(skip_photog_cb, state_pm_skip)
        assert state_pm_skip.state == admin_handlers.PostMatsForm.stream_record_url
        
        # Question 2: skip stream
        await admin_handlers.process_skip_pm_stream(skip_stream_cb, state_pm_skip)
        assert state_pm_skip.state == admin_handlers.PostMatsForm.article_url
        
        # Question 3: skip article
        await admin_handlers.process_skip_pm_article(skip_art_cb, state_pm_skip)
        assert state_pm_skip.state == admin_handlers.PostMatsForm.presentations_url
        
        # Question 4: skip presentations
        await admin_handlers.process_skip_pm_presentations(skip_pres_cb, state_pm_skip)
        assert state_pm_skip.state == admin_handlers.PostMatsForm.other_materials_url
        
        # Question 5: finish
        await admin_handlers.process_finish_pm_adding_btn(finish_pm_cb, state_pm_skip)
        assert state_pm_skip.state is None # FSM finished
        
        # Verify photographer fields are now cleared
        async with async_session() as test_session:
            ev2 = await test_session.get(Event, 2)
            assert ev2.photographer_name is None
            assert ev2.photographer_url is None

        # ----------------------------------------------------
        # 9. Scheduler Reminders
        # ----------------------------------------------------
        print("Testing scheduler reminders check...")
        mock_bot = DummyBot()
        
        # Calculate exactly 24 hours from now
        target_dt = datetime.now() + timedelta(hours=24)
        event_act.date = target_dt.strftime("%d.%m.%Y")
        event_act.time = target_dt.strftime("%H:%M")
        session.add(event_act)
        await session.commit()

        # Run scheduler tick
        await scheduler.check_and_send_reminders(mock_bot)
        assert len(mock_bot.sent_messages) > 0
        chat_id, text, kb = mock_bot.sent_messages[0]
        assert chat_id == 123
        assert "напоминание о мероприятии" in text.lower()
        assert kb is not None
        button_texts = [btn.text for row in kb.inline_keyboard for btn in row]
        assert "Подробнее о мероприятии" in button_texts


        # ----------------------------------------------------
        # 10. Admin Archived Event Edit Flow
        # ----------------------------------------------------
        print("Testing admin archived event edit flow...")
        # Mark event_dis (Digital design stuff) as hidden by changing its date to past
        past_dt = datetime.now() - timedelta(days=5)
        event_dis.hide_date = past_dt.strftime("%d.%m.%Y")
        event_dis.hide_time = "12:00"
        session.add(event_dis)
        await session.commit()

        # Trigger "admin_edit_archive_tags"
        cb_archtags = DummyCallbackQuery("admin_edit_archive_tags", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_edit_archive_tags(cb_archtags)
        # Verify message popped up with options
        assert "Выберите тему" in msg.sent_messages[-1][0]

        # Trigger tag select (index 1 is Digital_Дизайн)
        cb_archtag_sel = DummyCallbackQuery("admin_archtag_1", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_archtag_select(cb_archtag_sel)
        # Verify event 1 is listed
        arch_kb = msg.sent_messages[-1][1]
        arch_btn_texts = [btn.text for row in arch_kb.inline_keyboard for btn in row]
        assert "Digital design stuff" in arch_btn_texts
        print("Admin archived event edit flow test PASSED!")

        # Delete test user 456 to prevent notification test interference
        u456 = await session.get(User, 456)
        if u456:
            await session.delete(u456)
            await session.commit()

        # ----------------------------------------------------
        # 11. Event and Raffle Creation Notifications
        # ----------------------------------------------------
        print("Testing event and raffle creation notifications...")
        mock_bot_creation = DummyBot()
        
        # Test Event Notification
        # user 123 has "Наука" enabled, should receive event_act notification
        await admin_handlers.send_event_creation_notifications(mock_bot_creation, event_act)
        assert len(mock_bot_creation.sent_messages) > 0
        assert mock_bot_creation.sent_messages[0][0] == 123
        assert "появилось мероприятие" in mock_bot_creation.sent_messages[0][1].lower()

        # Test Raffle Notification
        mock_bot_creation2 = DummyBot()
        # raffle is linked to event 100 which has tag "Digital_Дизайн"
        # user 123 has "Digital_Дизайн" disabled, should not receive it
        raffle_test = Raffle(
            id=300,
            title="Test Raffle Notification",
            description="Awesome description",
            is_active=1,
            event_id=1
        )
        await admin_handlers.send_raffle_creation_notifications(mock_bot_creation2, raffle_test)
        assert len(mock_bot_creation2.sent_messages) == 0

        # Change user 123 preferences to enable "Digital_Дизайн"
        user.tags_preferences = {k: True for k in config.DEFAULT_TAGS}
        session.add(user)
        await session.commit()

        await admin_handlers.send_raffle_creation_notifications(mock_bot_creation2, raffle_test)
        assert len(mock_bot_creation2.sent_messages) > 0
        assert "появился розыгрыш" in mock_bot_creation2.sent_messages[0][1].lower()
        print("Event and raffle creation notifications test PASSED!")

        # ----------------------------------------------------
        # 12. Admin Registrations Excel Export
        # ----------------------------------------------------
        print("Testing admin registrations Excel export...")
        # Trigger export select category
        cb_export = DummyCallbackQuery("admin_export_registrations", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_export_registrations(cb_export)
        assert "Выберите категорию" in msg.sent_messages[-1][0]

        # Trigger active events selection
        cb_export_act = DummyCallbackQuery("admin_export_select_active", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_export_select_active(cb_export_act)
        assert "Выберите активное мероприятие" in msg.sent_messages[-1][0]

        # Seed registrations with "думаю" and "очно" statuses to test sorting
        async with async_session() as test_session:
            # Let's clean up existing registrations for event 2
            await test_session.execute(Registration.__table__.delete().where(Registration.event_id == 2))
            
            # Add user 456
            u456 = await test_session.get(User, 456)
            if not u456:
                u456 = User(telegram_id=456, username="dummy456", full_name="Алексеев Алексей", is_registered=True, notification_preferences={}, tags_preferences={})
                test_session.add(u456)
            
            # Add "думаю" registration (first, so naturally it would be first without sorting)
            r_think = Registration(user_id=123, event_id=2, status="думаю", registration_date="25.06.2026")
            test_session.add(r_think)
            
            # Add "очно" registration (second)
            r_active = Registration(user_id=456, event_id=2, status="очно", registration_date="25.06.2026")
            test_session.add(r_active)
            await test_session.commit()

        # Trigger actual Excel file generation for event 2
        cb_export_ev = DummyCallbackQuery("admin_export_event_2", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_export_event(cb_export_ev)
        
        # Verify the file was generated and sent
        sent_doc_msg = msg.sent_messages[-1]
        assert "Регистраций:" in sent_doc_msg[0]
        # The document is a BufferedInputFile
        from aiogram.types import BufferedInputFile
        assert isinstance(sent_doc_msg[1], BufferedInputFile)
        assert sent_doc_msg[1].filename.startswith("registrations_2_")
        assert sent_doc_msg[2] is not None  # reply_markup check
        
        # Let's inspect the excel data bytes to make sure they are a valid workbook
        import io
        import openpyxl
        file_bytes = sent_doc_msg[1].data
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        assert ws.title == "Регистрации"
        row1 = [cell.value for cell in ws[1]]
        assert row1 == ["ФИО", "Электронная почта", "Телефон", "Откуда вы узнали о событии?", "Вы планируете быть очно или удалённо?", "Дата", "TG ник"]
        
        # Row 2 should be the active one ("очно") because "думаю" is sorted to the end
        row2 = [cell.value for cell in ws[2]]
        assert row2[0] == "Алексеев Алексей"
        assert row2[4] == "очно"
        assert row2[6] == "@dummy456"

        # Row 3 should be the thinking one ("думаю")
        row3 = [cell.value for cell in ws[3]]
        assert row3[0] == "Кузнецов Пётр"
        assert row3[4] == "думаю"
        assert row3[6] == "@john_doe"

        print("Admin registrations Excel export test PASSED!")


        # ----------------------------------------------------
        # 13. User Data Deletion (/delete Command)
        # ----------------------------------------------------
        print("Testing user registration deletion via /delete...")
        # Verify user 123 is registered beforehand
        u123 = await session.get(User, 123)
        assert u123.is_registered is True
        assert u123.full_name == "Кузнецов Пётр"

        # Check registrations exist for user 123
        reg123 = await session.get(Registration, (123, 2))
        assert reg123 is not None

        # Execute /delete
        delete_state = DummyState()
        await handlers.cmd_delete(DummyMessage(123, "john_doe"), delete_state)

        # Clear identity map cache so we fetch fresh state from DB
        session.expire_all()

        # Refresh database session and verify user fields are reset
        await session.refresh(u123)
        assert u123.is_registered is False
        assert u123.full_name is None
        assert u123.email is None
        assert u123.phone == "-"
        # Also ensure FSM registration flow is restarted for the user
        assert delete_state.state == handlers.RegistrationStates.accept_agreement

        # Ensure registrations are deleted
        reg123_check = await session.get(Registration, (123, 2))
        assert reg123_check is None

        print("User data deletion via /delete test PASSED!")

        # ----------------------------------------------------
        # 14. Admin Default Event Address Button
        # ----------------------------------------------------
        print("Testing admin default event address option...")
        address_state = DummyState()
        await address_state.set_state(admin_handlers.EventForm.address)
        
        from admin_keyboards import get_address_keyboard
        kb = get_address_keyboard(show_back=True, back_callback="back_address")
        button_texts = [btn.text for row in kb.inline_keyboard for btn in row]
        assert "Вставить адрес Хаба" in button_texts
        
        msg_address = DummyMessage(999, "ASaavedraA")
        cb_address = DummyCallbackQuery("admin_event_default_address", 999, "ASaavedraA", msg_address)
        await admin_handlers.process_add_event_default_address(cb_address, address_state)
        
        data = await address_state.get_data()
        assert data.get("address") == "Москва, Пантелеевская 53"
        assert address_state.state == admin_handlers.EventForm.description
        assert "Вопрос 6: Основной текст" in msg_address.sent_messages[-1][0]
        print("Admin default event address option test PASSED!")

        # ----------------------------------------------------
        # 15. Admin Post-Materials Deletion Flow
        # ----------------------------------------------------
        print("Testing admin post-materials deletion flow...")
        # Populate event 2 with a photo url to test deletion
        async with async_session() as test_session:
            ev2 = await test_session.get(Event, 2)
            ev2.photos_url = "https://photos.url/123"
            test_session.add(ev2)
            await test_session.commit()

        # Verify event 2 has photos_url first
        async with async_session() as test_session:
            ev2 = await test_session.get(Event, 2)
            assert ev2.photos_url == "https://photos.url/123"

        msg_pm_del = DummyMessage(999, "ASaavedraA")
        # Trigger confirm step (admin_pm_del_2)
        cb_pm_del_start = DummyCallbackQuery("admin_pm_del_2", 999, "ASaavedraA", msg_pm_del)
        await admin_handlers.process_admin_pm_del_confirm_start(cb_pm_del_start)
        assert "Вы уверены? Это действие нельзя отменить!" in msg_pm_del.sent_messages[-1][0]
        
        # Trigger final deletion step (admin_pm_del_confirm_2)
        cb_pm_del_final = DummyCallbackQuery("admin_pm_del_confirm_2", 999, "ASaavedraA", msg_pm_del)
        await admin_handlers.process_admin_pm_del_final(cb_pm_del_final)
        assert "Пост-материалы удалены!" in msg_pm_del.sent_messages[-1][0]
        
        # Verify database fields are updated to None
        async with async_session() as test_session:
            ev2_updated = await test_session.get(Event, 2)
            assert ev2_updated.photos_url is None
        print("Admin post-materials deletion flow test PASSED!")

        # ----------------------------------------------------
        # 16. Stream URL Notifications for Remote Participants
        # ----------------------------------------------------
        print("Testing stream URL update notifications...")
        
        async with async_session() as test_session:
            existing_reg = await test_session.get(Registration, (123, 2))
            if existing_reg:
                await test_session.delete(existing_reg)
                await test_session.commit()
                
            remote_reg = Registration(
                user_id=123,
                event_id=2,
                status="удаленно",
                registration_date="25.06.2026"
            )
            test_session.add(remote_reg)
            
            ev2 = await test_session.get(Event, 2)
            ev2.stream_url = None
            ev2.images = ["AgACAgIAAxkBAAIB"]
            test_session.add(ev2)
            await test_session.commit()
            
        edit_state = DummyState()
        await edit_state.set_state(admin_handlers.EditEventForm.value)
        await edit_state.update_data(event_id=2, section="streamurl")
        
        mock_bot_stream = DummyBot()
        edit_msg = DummyMessage(999, "ASaavedraA", "https://youtube.com/live_stream")
        edit_msg.bot = mock_bot_stream
        
        await admin_handlers.process_edit_event_value(edit_msg, edit_state)
        
        await asyncio.sleep(0.1)
        
        async with async_session() as test_session:
            ev2_updated = await test_session.get(Event, 2)
            assert ev2_updated.stream_url == "https://youtube.com/live_stream"
            
        assert len(mock_bot_stream.sent_messages) > 0
        
        broadcast_msg = None
        for chat_id, content, kb in mock_bot_stream.sent_messages:
            if chat_id == 123:
                broadcast_msg = content
                break
                
        assert broadcast_msg is not None
        assert "Ссылка на трансляцию мероприятия уже доступна!" in broadcast_msg
        assert "Вы зарегистрировались на трансляцию мероприятия" in broadcast_msg
        assert "Трансляция" in broadcast_msg
        assert "youtube.com/live_stream" in broadcast_msg
        print("Stream URL update notifications test PASSED!")

        # ----------------------------------------------------
        # 17. User Feedback Flow & Admin Navigator
        # ----------------------------------------------------
        print("Testing feedback flow & admin navigator...")
        
        # 17.1 Admin vs User Main Menu Layout Test
        kb_admin = get_main_menu_keyboard(is_admin=True, raffle_count=0)
        kb_user = get_main_menu_keyboard(is_admin=False, raffle_count=0)
        
        # Verify admins do NOT see Feedback button
        admin_buttons = [btn.text for row in kb_admin.inline_keyboard for btn in row]
        assert "Обратная связь" not in admin_buttons
        
        # Verify users DO see Feedback button without emoji
        user_buttons = [btn.text for row in kb_user.inline_keyboard for btn in row]
        assert "Обратная связь" in user_buttons
        assert "✍️" not in "".join(user_buttons)
        
        # 17.2 User starting and cancelling feedback
        feedback_state = DummyState()
        msg_fb = DummyMessage(123, "john_doe")
        msg_fb.bot = DummyBot()
        cb_fb_start = DummyCallbackQuery("btn_feedback", 123, "john_doe", msg_fb)
        
        await handlers.process_feedback_start(cb_fb_start, feedback_state)
        assert feedback_state.state == handlers.FeedbackStates.waiting_for_message
        assert "Обратная связь" in msg_fb.sent_messages[-1][0]
        
        cb_fb_cancel = DummyCallbackQuery("cancel_feedback", 123, "john_doe", msg_fb)
        await handlers.process_feedback_cancel(cb_fb_cancel, feedback_state)
        assert feedback_state.state is None
        assert "Мероприятия по вашим предпочтениям" in msg_fb.sent_messages[-1][0]
        
        # 17.3 User submitting feedback (checking DB storage and admin notifications)
        # Clear database feedbacks to start fresh
        async with async_session() as test_session:
            await test_session.execute(delete(FeedbackMessage))
            await test_session.commit()
            
        await feedback_state.set_state(handlers.FeedbackStates.waiting_for_message)
        fb_msg = DummyMessage(123, "john_doe", "Привет, это фидбек!")
        fb_msg.bot = DummyBot()
        
        # Ensure super admin 999 is in database
        async with async_session() as test_session:
            admin_user = await test_session.get(User, 999)
            if not admin_user:
                admin_user = User(telegram_id=999, username="ASaavedraA", is_registered=True)
                test_session.add(admin_user)
                await test_session.commit()
        
        await handlers.process_feedback_message(fb_msg, feedback_state)
        assert feedback_state.state is None
        assert "успешно отправлено" in fb_msg.sent_messages[-2][0]
        
        # Verify it was stored in Database
        async with async_session() as test_session:
            stored_feedbacks = (await test_session.execute(select(FeedbackMessage))).scalars().all()
            assert len(stored_feedbacks) == 1
            assert stored_feedbacks[0].text == "Привет, это фидбек!"
            assert stored_feedbacks[0].user_id == 123
        
        # Verify notification sent to admin includes Back ("Назад") button
        admin_received = False
        for chat_id, content, kb in fb_msg.bot.sent_messages:
            if chat_id == 999 and "Новое сообщение обратной связи" in content:
                assert "Привет, это фидбек!" in content
                # Verify "Назад" button exists in reply markup
                kb_buttons = [btn.text for row in kb.inline_keyboard for btn in row]
                assert "Назад" in kb_buttons
                admin_received = True
                break
        assert admin_received is True
        
        # 17.4 Admin Feedback Viewer Empty State
        # Let's temporarily delete all feedbacks to test empty state
        async with async_session() as test_session:
            await test_session.execute(delete(FeedbackMessage))
            await test_session.commit()
            
        admin_msg = DummyMessage(999, "ASaavedraA")
        cb_view_empty = DummyCallbackQuery("admin_view_feedback", 999, "ASaavedraA", admin_msg)
        view_state = DummyState()
        await admin_handlers.process_admin_view_feedback(cb_view_empty, view_state)
        assert "Сообщений обратной связи пока нет" in admin_msg.sent_messages[-1][0]
        assert view_state.state is None
        
        # 17.5 Admin Feedback Viewer Navigation Loop
        # Insert 3 test feedbacks
        async with async_session() as test_session:
            f1 = FeedbackMessage(user_id=101, full_name="User One", text="Фидбек 1", created_at="2026-06-30 10:00:00")
            f2 = FeedbackMessage(user_id=102, full_name="User Two", text="Фидбек 2", created_at="2026-06-30 10:01:00")
            f3 = FeedbackMessage(user_id=103, full_name="User Three", text="Фидбек 3", created_at="2026-06-30 10:02:00")
            test_session.add_all([f1, f2, f3])
            await test_session.commit()
            
        cb_view = DummyCallbackQuery("admin_view_feedback", 999, "ASaavedraA", admin_msg)
        await admin_handlers.process_admin_view_feedback(cb_view, view_state)
        # Should display the LATEST (newest) feedback first: "Фидбек 3" (index 2 out of 3)
        assert view_state.state == admin_handlers.AdminFeedbackStates.viewing
        data = await view_state.get_data()
        assert data.get("current_feedback_index") == 2
        assert "Фидбек 3" in admin_msg.sent_messages[-1][0]
        
        # Click "Дальше" (next) -> should loop around to index 0: "Фидбек 1"
        cb_next = DummyCallbackQuery("feedback_nav_next", 999, "ASaavedraA", admin_msg)
        await admin_handlers.process_feedback_nav_next(cb_next, view_state)
        data = await view_state.get_data()
        assert data.get("current_feedback_index") == 0
        assert "Фидбек 1" in admin_msg.sent_messages[-1][0]
        
        # Click "Назад" (prev) -> should loop around to index 2: "Фидбек 3"
        cb_prev = DummyCallbackQuery("feedback_nav_prev", 999, "ASaavedraA", admin_msg)
        await admin_handlers.process_feedback_nav_prev(cb_prev, view_state)
        data = await view_state.get_data()
        assert data.get("current_feedback_index") == 2
        assert "Фидбек 3" in admin_msg.sent_messages[-1][0]
        
        # Click "Назад" (prev) again -> should go to index 1: "Фидбек 2"
        await admin_handlers.process_feedback_nav_prev(cb_prev, view_state)
        data = await view_state.get_data()
        assert data.get("current_feedback_index") == 1
        assert "Фидбек 2" in admin_msg.sent_messages[-1][0]
        
        print("Feedback flow & admin navigator test PASSED!")

        # ----------------------------------------------------
        # 18. Preference Tags Prompt & Unread Post-Materials Indicators
        # ----------------------------------------------------
        print("Testing preference tags prompt & unread post-materials indicators...")
        
        # 18.1 Check update preferences description text
        pref_state = DummyState()
        pref_msg = DummyMessage(123, "john_doe")
        pref_cb = DummyCallbackQuery("pref_tags", 123, "john_doe", pref_msg)
        await handlers.process_pref_tags(pref_cb, pref_state)
        
        prompt_sent = pref_msg.sent_messages[-1][0]
        assert "Прокликайте темы мероприятий, о которых хотите получать уведомления." in prompt_sent
        assert "✅ — уведомления включены" in prompt_sent
        assert "❌ — уведомления выключены" in prompt_sent
        
        # 18.2 Check unread post-materials logic
        # Setup: Clear ReadPostMaterials and ensure user 123 has "Наука" enabled, "Digital_Дизайн" disabled
        async with async_session() as test_session:
            await test_session.execute(delete(ReadPostMaterials))
            
            # Setup User preferences
            user = await test_session.get(User, 123)
            user.tags_preferences = {"Наука": True, "Digital_Дизайн": False}
            test_session.add(user)
            
            # Clear events, add 1 event under "Наука" with post materials, 1 event under "Digital_Дизайн" with post materials
            await test_session.execute(delete(Event))
            
            ev_science = Event(
                id=10,
                title="Научный Семинар",
                date="12.07.2026",
                time="18:00",
                address="Пантелеевская, 53",
                photos_url="http://science.url",
                tags=["Наука"]
            )
            ev_design = Event(
                id=11,
                title="Дизайн Студия",
                date="13.07.2026",
                time="19:00",
                address="Пантелеевская, 53",
                photos_url="http://design.url",
                tags=["Digital_Дизайн"]
            )
            test_session.add_all([ev_science, ev_design])
            await test_session.commit()
            
        # Call process_archive_menu
        archive_msg = DummyMessage(123, "john_doe")
        archive_cb = DummyCallbackQuery("btn_archive", 123, "john_doe", archive_msg)
        await handlers.process_archive_menu(archive_cb, DummyState())
        
        # Check that markup has "🔔 Наука" (enabled, unread)
        # but does NOT have "🔔 Digital_Дизайн" (disabled, unread)
        tags_kb = archive_msg.sent_messages[-1][1]
        buttons = [btn.text for row in tags_kb.inline_keyboard for btn in row]
        assert "🔔 Наука" in buttons
        assert "Digital_Дизайн" in buttons
        assert "🔔 Digital_Дизайн" not in buttons
        
        # Go to tag selection list
        tag_select_msg = DummyMessage(123, "john_doe")
        # "Наука" is index 0 in config.DEFAULT_TAGS
        science_idx = config.DEFAULT_TAGS.index("Наука")
        tag_cb = DummyCallbackQuery(f"arch_tag_{science_idx}", 123, "john_doe", tag_select_msg)
        await handlers.process_archive_tag_selection(tag_cb)
        
        # Check event list markup has "🔔 Научный Семинар"
        events_kb = tag_select_msg.sent_messages[-1][1]
        event_buttons = [btn.text for row in events_kb.inline_keyboard for btn in row]
        assert "🔔 Научный Семинар" in event_buttons
        
        # Transition to event detail (viewing it)
        detail_msg = DummyMessage(123, "john_doe")
        event_cb = DummyCallbackQuery("arch_event_10", 123, "john_doe", detail_msg)
        await handlers.process_archive_event_detail(event_cb)
        
        # Verify read entry is created in database
        async with async_session() as test_session:
            read_status = await test_session.get(ReadPostMaterials, (123, 10))
            assert read_status is not None
            
        # Re-check tag list
        tag_select_msg_2 = DummyMessage(123, "john_doe")
        await handlers.process_archive_tag_selection(DummyCallbackQuery(f"arch_tag_{science_idx}", 123, "john_doe", tag_select_msg_2))
        events_kb_2 = tag_select_msg_2.sent_messages[-1][1]
        event_buttons_2 = [btn.text for row in events_kb_2.inline_keyboard for btn in row]
        # Fire emoji should be gone!
        assert "Научный Семинар" in event_buttons_2
        assert "🔔 Научный Семинар" not in event_buttons_2
        
        # Re-check categories menu
        archive_msg_2 = DummyMessage(123, "john_doe")
        await handlers.process_archive_menu(DummyCallbackQuery("btn_archive", 123, "john_doe", archive_msg_2), DummyState())
        tags_kb_2 = archive_msg_2.sent_messages[-1][1]
        buttons_2 = [btn.text for row in tags_kb_2.inline_keyboard for btn in row]
        # Fire emoji should be gone from Наука!
        assert "Наука" in buttons_2
        assert "🔔 Наука" not in buttons_2
        
        print("Preference tags prompt & unread post-materials indicators test PASSED!")

        # ----------------------------------------------------
        # 19. Admin Statistics & Buttons Registration Counts
        # ----------------------------------------------------
        print("Testing admin welcome stats & buttons registration counts...")
        
        # We need to set some users' created_at dates to check statistics
        async with async_session() as test_session:
            # Set creation times for test users
            u123_db = await test_session.get(User, 123)
            if u123_db:
                u123_db.created_at = "2026-07-09 10:00:00"
                test_session.add(u123_db)
            
            # Add registrations for event 10 to ensure count is 2
            await test_session.execute(delete(Registration))
            reg1 = Registration(user_id=123, event_id=10, status="очно", registration_date="09.07.2026")
            reg2 = Registration(user_id=999, event_id=10, status="думаю", registration_date="09.07.2026")
            test_session.add_all([reg1, reg2])
            await test_session.commit()
            
        admin_cb_stats = DummyCallbackQuery("btn_admin", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_menu(admin_cb_stats, state)
        
        admin_menu_msg = msg.sent_messages[-1][0]
        assert "Уникальных пользователей бота:" in admin_menu_msg
        assert "Новых пользователей за сутки:" in admin_menu_msg
        assert "Новых пользователей за неделю:" in admin_menu_msg
        
        # Test active events selection list shows registration count on buttons
        # Event 2 has registrations (Aлексеев Алексей and Кузнецов Пётр registered) -> count should be 2.
        active_export_msg = DummyMessage(999, "ASaavedraA")
        active_export_cb = DummyCallbackQuery("admin_export_select_active", 999, "ASaavedraA", active_export_msg)
        await admin_handlers.process_admin_export_select_active(active_export_cb)
        
        active_kb = active_export_msg.sent_messages[-1][1]
        active_buttons = [btn.text for row in active_kb.inline_keyboard for btn in row]
        # Active events keyboard has event 2 (Слабое звено...) which has 2 registrations.
        # It also has other events. Let's make sure the count matches "(2) "
        assert any(btn.startswith("(2) ") for btn in active_buttons)
        
        print("Admin welcome stats & buttons registration counts test PASSED!")

        # ----------------------------------------------------
        # 20. Partner Events & Date Range Support
        # ----------------------------------------------------
        print("Testing partner events and date range support...")
        
        # 20.1 Check date range formatting
        assert config.format_display_date("16.07.2026-17.07.2026") == "16 июля 2026 — 17 июля 2026"
        assert config.format_display_date("12.06.2026") == "12 июня 2026"
        assert config.format_partner_date("16.07.2026-17.07.2026") == "16 июля-17 июля 2026"
        assert config.format_partner_date("16.07.2026-17.07.2027") == "16 июля 2026-17 июля 2027"
        
        # 20.2 Create a PartnerEvent via FSM
        p_state = DummyState()
        # Step 1: title
        p_cb = DummyCallbackQuery("admin_add_partner_event", 999, "ASaavedraA", msg)
        await admin_handlers.process_admin_add_partner_event(p_cb, p_state)
        assert p_state.state == admin_handlers.PartnerEventForm.title
        
        # Step 2: date
        await admin_handlers.process_add_partner_title(DummyMessage(999, "ASaavedraA", "Партнерский Ивент"), p_state)
        assert p_state.state == admin_handlers.PartnerEventForm.date
        
        # Try invalid date format
        bad_date_msg = DummyMessage(999, "ASaavedraA", "invalid-date")
        await admin_handlers.process_add_partner_date(bad_date_msg, p_state)
        assert p_state.state == admin_handlers.PartnerEventForm.date # stays on date
        
        # Valid date range
        await admin_handlers.process_add_partner_date(DummyMessage(999, "ASaavedraA", "16.07.2026-17.07.2026"), p_state)
        assert p_state.state == admin_handlers.PartnerEventForm.description
        
        # Step 3: description
        await admin_handlers.process_add_partner_description(DummyMessage(999, "ASaavedraA", "Описание партнера"), p_state)
        assert p_state.state == admin_handlers.PartnerEventForm.link
        
        # Step 4: link
        await admin_handlers.process_add_partner_link(DummyMessage(999, "ASaavedraA", "https://partner.com"), p_state)
        assert p_state.state is None # FSM finished
        
        # Verify it is in database
        async with async_session() as test_session:
            pevents = (await test_session.execute(select(PartnerEvent))).scalars().all()
            assert len(pevents) == 1
            assert pevents[0].title == "Партнерский Ивент"
            assert pevents[0].date == "16.07.2026-17.07.2026"
            
        # 20.3 Test count on button (should show count of 1)
        user_msg = DummyMessage(123, "john_doe")
        user_cb = DummyCallbackQuery("btn_events_info", 123, "john_doe", user_msg)
        await handlers.process_events_info(user_cb, DummyState())
        
        user_kb = user_msg.sent_messages[-1][1]
        user_buttons = [btn.text for row in user_kb.inline_keyboard for btn in row]
        assert "Мероприятия партнеров (1)" in user_buttons
        
        # 20.4 Test compilation rendering, sorting & back button
        comp_msg = DummyMessage(123, "john_doe")
        comp_cb = DummyCallbackQuery("btn_partners_events", 123, "john_doe", comp_msg)
        await handlers.process_partners_events(comp_cb)
        
        comp_text = comp_msg.sent_messages[-1][0]
        assert "Мероприятия партнёров Хаба:" in comp_text
        assert "▪️ <b>16 июля-17 июля 2026 | Партнерский Ивент</b>" in comp_text
        assert "Описание партнера" in comp_text
        assert "→ Подробности" in comp_text
        
        comp_kb = comp_msg.sent_messages[-1][1]
        comp_btn_texts = [btn.text for row in comp_kb.inline_keyboard for btn in row]
        assert "Назад" in comp_btn_texts
        assert "← К списку" in comp_btn_texts
        
        # 20.5 Test deletion flow
        del_list_msg = DummyMessage(999, "ASaavedraA")
        del_list_cb = DummyCallbackQuery("admin_del_partner_event_list", 999, "ASaavedraA", del_list_msg)
        await admin_handlers.process_admin_del_partner_event_list(del_list_cb)
        
        del_kb = del_list_msg.sent_messages[-1][1]
        del_buttons = [btn.text for row in del_kb.inline_keyboard for btn in row]
        assert "Партнерский Ивент" in del_buttons
        
        # Perform deletion
        del_confirm_cb = DummyCallbackQuery(f"admin_del_pevent_{pevents[0].id}", 999, "ASaavedraA", DummyMessage(999, "ASaavedraA"))
        await admin_handlers.process_admin_del_pevent(del_confirm_cb)
        
        # Verify deleted
        async with async_session() as test_session:
            pevents_after = (await test_session.execute(select(PartnerEvent))).scalars().all()
            assert len(pevents_after) == 0
            
        print("Partner events and date range support test PASSED!")

        # ----------------------------------------------------
        # 21. Horizontal Navigation Buttons
        # ----------------------------------------------------
        print("Testing horizontal navigation buttons...")
        
        # 21.1 Active event detail keyboard
        kb_active_det = get_event_detail_keyboard(2)
        last_row_active = kb_active_det.inline_keyboard[-1]
        assert len(last_row_active) == 2
        assert last_row_active[0].text == "Назад" and last_row_active[0].callback_data == "btn_events_info"
        assert last_row_active[1].text == "← К списку" and last_row_active[1].callback_data == "back_to_main"
        
        # 21.2 Archive event detail keyboard
        kb_arch_det = get_archive_detail_keyboard(1)
        last_row_arch = kb_arch_det.inline_keyboard[-1]
        assert len(last_row_arch) == 2
        assert last_row_arch[0].text == "Назад" and last_row_arch[0].callback_data == "arch_tag_1"
        assert last_row_arch[1].text == "← К списку" and last_row_arch[1].callback_data == "back_to_main"
        
        # 21.3 Admin active events export selection keyboard
        kb_ex_active = get_admin_export_events_keyboard([])
        last_row_ex_active = kb_ex_active.inline_keyboard[-1]
        assert len(last_row_ex_active) == 2
        assert last_row_ex_active[0].text == "Назад" and last_row_ex_active[0].callback_data == "admin_export_registrations"
        assert last_row_ex_active[1].text == "← К списку" and last_row_ex_active[1].callback_data == "back_to_main"
        
        # 21.4 Admin archive events export selection keyboard
        kb_ex_arch = get_admin_export_archive_events_keyboard([], 1)
        last_row_ex_arch = kb_ex_arch.inline_keyboard[-1]
        assert len(last_row_ex_arch) == 2
        assert last_row_ex_arch[0].text == "Назад" and last_row_ex_arch[0].callback_data == "admin_export_select_archive"
        assert last_row_ex_arch[1].text == "← К списку" and last_row_ex_arch[1].callback_data == "back_to_main"
        
        # 21.5 Partners events list keyboard
        kb_partners = get_partners_events_keyboard()
        last_row_partners = kb_partners.inline_keyboard[-1]
        assert len(last_row_partners) == 2
        assert last_row_partners[0].text == "Назад" and last_row_partners[0].callback_data == "btn_events_info"
        assert last_row_partners[1].text == "← К списку" and last_row_partners[1].callback_data == "back_to_main"
        
        # 21.6 Admin events menu keyboard
        kb_ad_ev = get_admin_events_keyboard()
        last_row_ad_ev = kb_ad_ev.inline_keyboard[-1]
        assert len(last_row_ad_ev) == 2
        assert last_row_ad_ev[0].text == "Назад" and last_row_ad_ev[0].callback_data == "btn_admin"
        assert last_row_ad_ev[1].text == "← К списку" and last_row_ad_ev[1].callback_data == "back_to_main"
        
        # 21.7 Archive tags selection keyboard
        kb_arch_tags = get_archive_tags_keyboard([])
        last_row_arch_tags = kb_arch_tags.inline_keyboard[-1]
        assert len(last_row_arch_tags) == 2
        assert last_row_arch_tags[0].text == "Назад" and last_row_arch_tags[0].callback_data == "back_to_main"
        assert last_row_arch_tags[1].text == "← К списку" and last_row_arch_tags[1].callback_data == "back_to_main"
        
        # 21.8 Archive events list keyboard
        kb_arch_evs = get_archive_events_keyboard([])
        last_row_arch_evs = kb_arch_evs.inline_keyboard[-1]
        assert len(last_row_arch_evs) == 2
        assert last_row_arch_evs[0].text == "Назад" and last_row_arch_evs[0].callback_data == "btn_archive"
        assert last_row_arch_evs[1].text == "← К списку" and last_row_arch_evs[1].callback_data == "back_to_main"
        
        print("Horizontal navigation buttons test PASSED!")








def main():
    print("Starting comprehensive Bot Test Suite...")
    try:
        asyncio.run(test_suite())
        print("\nALL BOT HANDLERS & LOGIC TESTED SUCCESSFULLY WITH ZERO ERRORS!")
        # If success, clear test_errors.log if it exists
        if os.path.exists("test_errors.log"):
            os.remove("test_errors.log")
    except Exception as e:
        error_msg = f"TEST SUITE FAILED WITH EXCEPTION:\n{traceback.format_exc()}"
        print(error_msg, file=sys.stderr)
        with open("test_errors.log", "w", encoding="utf-8") as f:
            f.write(error_msg)
        sys.exit(1)

if __name__ == "__main__":
    main()
